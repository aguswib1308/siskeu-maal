"""
Generator Laporan ZIS bulanan ke Yayasan MKU Pusat (LAZ Pusat).

Mengisi 6 dari 8 sheet template Excel resmi HQ (Saldo Rekening & Program
Kegiatan tetap diisi manual oleh admin, tidak disentuh modul ini) langsung
dari database transaksi -- lihat rencana di
C:\\Users\\asus\\.claude\\plans\\wild-percolating-taco.md utk detail pemetaan
sel & alasan tiap keputusan.

Batasan yang disengaja (data belum ada di sistem):
- Tunai/Transfer/Instant Payment: semua penerimaan dihitung sbg Tunai --
  sistem belum lacak channel transfer/QRIS per transaksi.
- CSR/Qurban/DSKL/Hibah sbg kategori penerimaan terpisah: selalu 0, krn COA
  belum py akun penerimaan khusus utk ini (semua masuk generik ke Infaq
  Bebas/Terikat). Tambah akunnya nanti kalau sudah ada transaksi nyata.
- Jumlah hewan qurban: tidak ada field, dibiarkan kosong.
- Penyaluran Zakat -> Bidang Program: lihat map_zakat_ke_bidang() di bawah,
  masih placeholder menunggu pola dari user.
"""
import os
import io
from openpyxl import load_workbook

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              'laporan_template', 'laz_pusat_template.xlsx')

# kode COA (leaf/parent) -> bidang program, dipakai bareng utk sheet Pentasyarufan
# (bagian Bidang Program) & Penerima Manfaat. '5.2.x' = infaq tidak terikat,
# '5.5.x' = infaq terikat (struktur paralel, lihat init_db.py).
BIDANG_MAP = [
    ('5.2.1', 'pendidikan'), ('5.5.1', 'pendidikan'),
    ('5.2.2', 'kesehatan'),  ('5.5.2', 'kesehatan'),
    ('5.2.3', 'ekonomi'),    ('5.5.3', 'ekonomi'),
    ('5.2.4', 'sosial'),     ('5.5.4', 'sosial'),
    ('5.2.5', 'sosial'),     ('5.5.5', 'sosial'),   # Bencana masuk Sosial/Kemanusiaan
    ('5.2.6', 'dakwah'),     ('5.5.6', 'dakwah'),
    ('5.2.7', 'qurban'),     ('5.5.7', 'qurban'),
]
# Bukan program -- transfer/operasional internal, dikecualikan dari Bidang & sub-tabel Infaq
EXCLUDE_DARI_BIDANG = {'5.2.6.05', '5.2.8'}  # Operasional Amil [OP], Hibah ke Dana Lain

ASNAF_KODE = {'5.1.1': 'fakir', '5.1.2': 'miskin', '5.1.3': 'amil', '5.1.4': 'muallaf',
              '5.1.5': 'riqab', '5.1.6': 'gharim', '5.1.7': 'fisabilillah', '5.1.8': 'ibnu_sabil'}

# kode akun Beban Dana Amil -> baris "Penggunaan Hak Amil" di sheet Dana Amil (1-9)
AMIL_BARIS_MAP = {
    '5.3.1': 1,       # Belanja Pegawai (Gaji)
    '5.3.3': 2,       # Biaya Publikasi Dan Dokumentasi
    '5.3.4': 4,       # Beban Administrasi Umum
    '5.3.2': 8,       # Penggunaan lain hak amil
    '5.2.6.05': 9,    # Disalurkan bersama dana pentasharufan (Operasional Amil [OP])
}

KOLOM_BULAN = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']  # index 0=Jan


def _bulan_of(tanggal):
    return int(tanggal[5:7])


def _match_prefix(kode, prefix):
    return kode == prefix or kode.startswith(prefix + '.')


def bidang_dari_kode(kode):
    if not kode or kode in EXCLUDE_DARI_BIDANG:
        return None
    for prefix, bidang in BIDANG_MAP:
        if _match_prefix(kode, prefix):
            return bidang
    return None


def map_zakat_ke_bidang(keterangan):
    """Placeholder -- menunggu pola dari user utk cocokkan penyaluran zakat (per-asnaf)
    ke bidang program (Pendidikan/Kesehatan/dst) dari teks `keterangan`. Sampai diisi,
    penyaluran zakat TIDAK masuk ke rekap Bidang Program (dipisah di tabel Asnaf sendiri,
    sesuai default yg dipilih user)."""
    return None


def _kosong_bulanan(**default_keys):
    return {m: dict(default_keys) for m in range(1, 13)}


def data_penghimpunan(conn, tahun):
    hasil = _kosong_bulanan(tunai=0, transfer=0, instant=0, zakat_maal=0, zakat_fitrah=0,
                             infaq_bebas=0, infaq_terikat=0, csr=0, qurban=0, dskl=0, hibah=0)
    rows = conn.execute("""
        SELECT t.tanggal, t.jumlah, c.kode, c.jenis_dana
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id = c.id
        WHERE t.jenis='masuk' AND strftime('%Y', t.tanggal)=?
    """, (str(tahun),)).fetchall()
    for r in rows:
        bucket = _bucket_penghimpunan(r['kode'], r['jenis_dana'])
        if not bucket:
            continue  # 'amil'/'wakaf' masuk -- di luar cakupan laporan ZIS ini, lihat batasan modul
        h = hasil[_bulan_of(r['tanggal'])]
        h['tunai'] += r['jumlah']          # semua channel dianggap tunai, lihat batasan modul
        h[bucket] += r['jumlah']
    return hasil


def _bucket_penghimpunan(kode, jenis_dana):
    if kode == '4.1.2':
        return 'zakat_fitrah'
    if jenis_dana == 'zakat':
        return 'zakat_maal'  # zakat maal + profesi/perniagaan/pertanian/fidyah, lihat batasan modul
    if jenis_dana == 'infak_tidak_terikat':
        return 'infaq_bebas'
    if jenis_dana == 'infak_terikat':
        return 'infaq_terikat'
    return None  # 'amil'/'wakaf' -- bukan bagian laporan ZIS ini


def jumlah_donatur(conn, tahun):
    """'Jumlah Donatur' di template resmi ternyata = jumlah SETORAN/transaksi, bukan jumlah
    donatur unik -- dicek: donatur_id kosong (NULL) di 100% data historis (setoran kotak
    infaq/kencleng, bukan per-nama), dan COUNT(*) transaksi per kategori per bulan cocok
    persis/nyaris persis dgn angka di laporan lama (mis. Zakat Maal Jan: 2 vs 2). Jadi
    dihitung sbg jumlah transaksi masuk, bukan COUNT(DISTINCT donatur_id)."""
    hasil = _kosong_bulanan(zakat_maal=0, zakat_fitrah=0, infaq_bebas=0, infaq_terikat=0,
                             csr=0, qurban=0, dskl=0, hibah=0)
    rows = conn.execute("""
        SELECT t.tanggal, c.kode, c.jenis_dana
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id = c.id
        WHERE t.jenis='masuk' AND strftime('%Y', t.tanggal)=?
    """, (str(tahun),)).fetchall()
    for r in rows:
        bucket = _bucket_penghimpunan(r['kode'], r['jenis_dana'])
        if bucket:
            hasil[_bulan_of(r['tanggal'])][bucket] += 1
    return hasil


def pentasyarufan(conn, tahun):
    bidang = _kosong_bulanan(pendidikan=0, kesehatan=0, sosial=0, ekonomi=0, dakwah=0, qurban=0)
    zakat_asnaf = _kosong_bulanan(fakir=0, miskin=0, amil=0, muallaf=0, riqab=0, gharim=0,
                                   fisabilillah=0, ibnu_sabil=0)
    infaq = _kosong_bulanan(infaq_bebas=0, infaq_terikat=0, csr=0, dskl=0, hibah=0, untuk_amil=0)
    qurban = _kosong_bulanan(dana=0, hewan=None)

    rows = conn.execute("""
        SELECT t.tanggal, t.jumlah, t.keterangan, c.kode, c.jenis_dana
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id = c.id
        WHERE t.jenis='keluar' AND strftime('%Y', t.tanggal)=?
    """, (str(tahun),)).fetchall()

    for r in rows:
        b = _bulan_of(r['tanggal'])
        kode = r['kode']
        jumlah = r['jumlah']
        if kode in ASNAF_KODE:
            zakat_asnaf[b][ASNAF_KODE[kode]] += jumlah
            bd = map_zakat_ke_bidang(r['keterangan'])
            if bd:
                bidang[b][bd] += jumlah
            continue
        if kode in EXCLUDE_DARI_BIDANG:
            continue  # transfer/operasional internal, bukan program (lihat batasan modul)
        bd = bidang_dari_kode(kode)
        if bd:
            bidang[b][bd] += jumlah
        if r['jenis_dana'] == 'infak_tidak_terikat':
            infaq[b]['infaq_bebas'] += jumlah
        elif r['jenis_dana'] == 'infak_terikat':
            infaq[b]['infaq_terikat'] += jumlah
        if kode == '5.5.7.01':  # Tebar Qurban [TQUR]
            qurban[b]['dana'] += jumlah

    amil_masuk = conn.execute("""
        SELECT strftime('%m', tanggal) bln, SUM(jumlah) total
        FROM transaksi WHERE jenis='masuk' AND jenis_dana='amil' AND strftime('%Y', tanggal)=?
        GROUP BY bln
    """, (str(tahun),)).fetchall()
    for r in amil_masuk:
        infaq[int(r['bln'])]['untuk_amil'] = r['total'] or 0

    return dict(bidang=bidang, zakat_asnaf=zakat_asnaf, infaq=infaq, qurban=qurban)


def penerima_manfaat(conn, tahun):
    hasil = _kosong_bulanan(pendidikan=0, kesehatan=0, sosial=0, ekonomi=0, dakwah=0, qurban=0)
    rows = conn.execute("""
        SELECT t.tanggal, t.jumlah_mustahik, c.kode
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id = c.id
        WHERE t.jenis='keluar' AND t.jumlah_mustahik IS NOT NULL AND strftime('%Y', t.tanggal)=?
    """, (str(tahun),)).fetchall()
    for r in rows:
        bd = bidang_dari_kode(r['kode'])
        if bd:
            hasil[_bulan_of(r['tanggal'])][bd] += r['jumlah_mustahik']
    return hasil


def dana_amil_penggunaan(conn, tahun):
    hasil = {m: {i: 0 for i in range(1, 10)} for m in range(1, 13)}
    rows = conn.execute("""
        SELECT t.tanggal, t.jumlah, c.kode
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id = c.id
        WHERE t.jenis='keluar' AND c.jenis_dana='amil' AND strftime('%Y', t.tanggal)=?
    """, (str(tahun),)).fetchall()
    for r in rows:
        baris = AMIL_BARIS_MAP.get(r['kode'])
        if baris:
            hasil[_bulan_of(r['tanggal'])][baris] += r['jumlah']
    return hasil


def hitung_kelengkapan_mustahik(conn, tahun, bulan_sampai):
    """Jumlah transaksi keluar (bukan Jurnal Umum) yg blm ada jumlah_mustahik-nya,
    per bulan -- dipakai utk peringatan di halaman laporan sebelum download."""
    row = conn.execute("""
        SELECT COUNT(*) n FROM transaksi
        WHERE jenis='keluar' AND jurnal_id IS NULL AND jumlah_mustahik IS NULL
          AND strftime('%Y', tanggal)=? AND CAST(strftime('%m', tanggal) AS INTEGER) <= ?
    """, (str(tahun), bulan_sampai)).fetchone()
    return row['n']


def isi_template(conn, tahun, bulan_sampai):
    """Isi template resmi HQ dgn data tahun `tahun`, bulan 1..bulan_sampai.
    Formula bawaan template (Penghitungan Dana Amil, sebagian Dana Amil) tidak
    disentuh -- otomatis kalkulasi ulang saat file dibuka di Excel. Return BytesIO."""
    wb = load_workbook(TEMPLATE_PATH)

    dp = data_penghimpunan(conn, tahun)
    jd = jumlah_donatur(conn, tahun)
    pt = pentasyarufan(conn, tahun)
    pm = penerima_manfaat(conn, tahun)
    da = dana_amil_penggunaan(conn, tahun)

    ws = wb['Data Penghimpunan']
    ws['A4'] = tahun  # perbaiki bug template asli (tertulis 2025 statis)
    for m in range(1, bulan_sampai + 1):
        row = 6 + m
        h = dp[m]
        ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}'] = h['tunai'], h['transfer'], h['instant']
        ws[f'F{row}'], ws[f'G{row}'] = h['zakat_maal'], h['zakat_fitrah']
        ws[f'H{row}'], ws[f'I{row}'] = h['infaq_bebas'], h['infaq_terikat']
        ws[f'J{row}'], ws[f'K{row}'], ws[f'L{row}'], ws[f'M{row}'] = h['csr'], h['qurban'], h['dskl'], h['hibah']

    ws = wb['Jumlah Donatur']
    ws['A4'] = tahun
    for m in range(1, bulan_sampai + 1):
        row = 6 + m
        h = jd[m]
        ws[f'B{row}'], ws[f'C{row}'] = h['zakat_maal'], h['zakat_fitrah']
        ws[f'D{row}'], ws[f'E{row}'] = h['infaq_bebas'], h['infaq_terikat']
        ws[f'F{row}'], ws[f'G{row}'], ws[f'H{row}'], ws[f'I{row}'] = h['csr'], h['qurban'], h['dskl'], h['hibah']

    ws = wb['Pentasyarufan']
    ws['A3'] = tahun
    for m in range(1, bulan_sampai + 1):
        b = pt['bidang'][m]
        row = 7 + m
        ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}'] = b['pendidikan'], b['kesehatan'], b['sosial']
        ws[f'E{row}'], ws[f'F{row}'], ws[f'G{row}'] = b['ekonomi'], b['dakwah'], b['qurban']

        z = pt['zakat_asnaf'][m]
        row = 25 + m
        ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}'] = z['fakir'], z['miskin'], z['amil']
        ws[f'E{row}'], ws[f'F{row}'], ws[f'G{row}'] = z['muallaf'], z['riqab'], z['gharim']
        ws[f'H{row}'], ws[f'I{row}'] = z['fisabilillah'], z['ibnu_sabil']

        i_ = pt['infaq'][m]
        row = 41 + m
        ws[f'B{row}'], ws[f'C{row}'] = i_['infaq_bebas'], i_['infaq_terikat']
        ws[f'D{row}'], ws[f'E{row}'], ws[f'F{row}'] = i_['csr'], i_['dskl'], i_['hibah']
        ws[f'G{row}'] = i_['untuk_amil']

        q = pt['qurban'][m]
        row = 58 + m
        ws[f'B{row}'] = q['dana']  # kolom C (jumlah hewan) dibiarkan kosong, lihat batasan modul

    ws = wb['Penerima Manfaat']
    ws['A4'] = tahun
    for m in range(1, bulan_sampai + 1):
        row = 6 + m
        h = pm[m]
        ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}'] = h['pendidikan'], h['kesehatan'], h['sosial']
        ws[f'E{row}'], ws[f'F{row}'], ws[f'G{row}'] = h['ekonomi'], h['dakwah'], h['qurban']

    ws = wb['Dana Amil']
    for m in range(1, 13):
        kol = KOLOM_BULAN[m - 1]
        src_row = 3 + m  # baris bulan ybs di sheet 'Penghitungan Dana Amil'
        ws[f'{kol}3'] = f"='Penghitungan Dana Amil'!C{src_row}"  # lengkapi pola formula (Jan-Mei kosong di template asli)
        ws[f'{kol}5'] = f"='Penghitungan Dana Amil'!K{src_row}"  # termasuk perbaikan D5 yg bolong di template asli
    for m in range(1, bulan_sampai + 1):
        kol = KOLOM_BULAN[m - 1]
        h = da[m]
        for baris in range(9, 18):
            ws[f'{kol}{baris}'] = h[baris - 8]

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
