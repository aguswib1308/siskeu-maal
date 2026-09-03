from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, send_file
import sqlite3, hashlib, os, re, json, calendar as cal_mod, io, shutil, glob as glob_mod, time
import uuid as uuid_mod
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from functools import wraps
import laz_pusat_report
from wa import kirim_wa, render_pesan

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY') or 'BmtMaal@2026!'
DB_PATH = os.path.join('data', 'keuangan.db')

VAPID_PUBLIC_KEY = os.environ.get('VAPID_PUBLIC_KEY', '')
VAPID_PRIVATE_KEY = os.environ.get('VAPID_PRIVATE_KEY', '')
VAPID_CLAIM_EMAIL = os.environ.get('VAPID_CLAIM_EMAIL', 'amalmuslim.bmt@gmail.com')

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=15)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=15000")
    return conn

def parse_jumlah(raw):
    """Parse input jumlah ('1.500.000' / '1500000' / 1500000). None jika tidak valid atau <= 0."""
    try:
        val = float(str(raw).replace('.', '').replace(',', '').strip())
    except (TypeError, ValueError):
        return None
    return val if val > 0 else None

def insert_transaksi(conn, tanggal, jenis, coa_id, donatur_id, penerima_id,
                     jumlah, keterangan, user_id, client_uuid=None,
                     nama_kegiatan=None, lokasi=None, jumlah_mustahik=None):
    """Insert transaksi dengan guard idempotensi client_uuid (anti double-submit).
    nama_kegiatan/lokasi/jumlah_mustahik opsional, dipakai utk penyaluran (keluar)
    spy bisa direkap di Laporan Kegiatan Penyaluran. Return (trx_id, duplikat)."""
    if client_uuid:
        existing = conn.execute("SELECT id FROM transaksi WHERE client_uuid=?",
                                (client_uuid,)).fetchone()
        if existing:
            return existing['id'], True
    jenis_dana = None
    if coa_id:
        row = conn.execute("SELECT jenis_dana FROM chart_of_accounts WHERE id=?", (coa_id,)).fetchone()
        if row: jenis_dana = row['jenis_dana']
    try:
        cur = conn.execute('''INSERT INTO transaksi
            (tanggal,jenis,jenis_dana,coa_id,donatur_id,penerima_id,jumlah,keterangan,user_id,client_uuid,
             nama_kegiatan,lokasi,jumlah_mustahik)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (tanggal, jenis, jenis_dana, coa_id, donatur_id or None, penerima_id or None,
             jumlah, keterangan, user_id, client_uuid or None,
             nama_kegiatan or None, lokasi or None, jumlah_mustahik or None))
    except sqlite3.IntegrityError:
        if client_uuid:
            existing = conn.execute("SELECT id FROM transaksi WHERE client_uuid=?",
                                    (client_uuid,)).fetchone()
            if existing:
                return existing['id'], True
        raise
    return cur.lastrowid, False

app.jinja_env.globals['new_form_uuid'] = lambda: uuid_mod.uuid4().hex

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            return redirect(url_for('marketing_dashboard'))
        return f(*args, **kwargs)
    return decorated

def get_tanggal_kerja():
    """Tanggal kerja admin (default hari ini) — dipakai sbg default tanggal saat
    input transaksi, supaya mudah memasukkan transaksi masa lalu berturut-turut
    tanpa ganti tanggal manual tiap kali."""
    return session.get('tanggal_kerja') or date.today().isoformat()

@app.context_processor
def inject_tanggal_kerja():
    tk = get_tanggal_kerja()
    return {'tanggal_kerja': tk, 'tanggal_kerja_is_today': tk == date.today().isoformat()}

@app.route('/admin/tanggal-kerja', methods=['POST'])
@admin_required
def set_tanggal_kerja():
    tgl = request.form.get('tanggal', '').strip()
    try:
        datetime.strptime(tgl, '%Y-%m-%d')
        session['tanggal_kerja'] = tgl
        flash(f'Tanggal kerja diatur ke {tgl}. Transaksi baru akan default ke tanggal ini.', 'success')
    except ValueError:
        flash('Tanggal tidak valid.', 'danger')
    return redirect(request.referrer or url_for('admin_dashboard'))

@app.route('/admin/tanggal-kerja/reset', methods=['POST'])
@admin_required
def reset_tanggal_kerja():
    session.pop('tanggal_kerja', None)
    flash('Tanggal kerja dikembalikan ke hari ini.', 'success')
    return redirect(request.referrer or url_for('admin_dashboard'))

def format_rupiah(angka):
    try:
        return f"Rp {int(angka):,}".replace(',', '.')
    except (TypeError, ValueError):
        return "Rp 0"

def terbilang(n):
    """Konversi angka ke kata-kata bahasa Indonesia (untuk slip)."""
    try:
        n = int(round(float(n)))
    except (TypeError, ValueError):
        return 'nol'
    if n == 0:
        return 'nol'
    angka = ['', 'satu', 'dua', 'tiga', 'empat', 'lima', 'enam', 'tujuh',
             'delapan', 'sembilan', 'sepuluh', 'sebelas']
    def helper(x):
        if x < 12:
            return angka[x]
        elif x < 20:
            return helper(x - 10) + ' belas'
        elif x < 100:
            return helper(x // 10) + ' puluh' + ((' ' + helper(x % 10)) if x % 10 else '')
        elif x < 200:
            return 'seratus' + ((' ' + helper(x - 100)) if x - 100 else '')
        elif x < 1000:
            return helper(x // 100) + ' ratus' + ((' ' + helper(x % 100)) if x % 100 else '')
        elif x < 2000:
            return 'seribu' + ((' ' + helper(x - 1000)) if x - 1000 else '')
        elif x < 1000000:
            return helper(x // 1000) + ' ribu' + ((' ' + helper(x % 1000)) if x % 1000 else '')
        elif x < 1000000000:
            return helper(x // 1000000) + ' juta' + ((' ' + helper(x % 1000000)) if x % 1000000 else '')
        elif x < 1000000000000:
            return helper(x // 1000000000) + ' miliar' + ((' ' + helper(x % 1000000000)) if x % 1000000000 else '')
        else:
            return helper(x // 1000000000000) + ' triliun' + ((' ' + helper(x % 1000000000000)) if x % 1000000000000 else '')
    return helper(n).strip()

app.jinja_env.filters['rupiah'] = format_rupiah

DANA_TYPES  = ['zakat', 'infak_tidak_terikat', 'infak_terikat', 'amil', 'wakaf']
# Kode akun GL (Kas & Saldo Dana) per jenis dana, utk ditampilkan di Neraca.
# infak_tidak_terikat & infak_terikat berbagi kode Kas/Saldo Infak yg sama krn
# keduanya blm punya akun Kas terpisah di COA (pemisahan hanya di sisi mutasi).
KODE_KAS_DANA   = {'zakat': '1.1.1', 'infak_tidak_terikat': '1.1.2', 'infak_terikat': '1.1.2', 'amil': '1.1.3', 'wakaf': '1.1.4'}
KODE_SALDO_DANA = {'zakat': '3.1',   'infak_tidak_terikat': '3.2',   'infak_terikat': '3.2',   'amil': '3.3',   'wakaf': '3.4'}
LABEL_DANA  = {'zakat':'Zakat','infak_tidak_terikat':'Infak/Sedekah Tidak Terikat',
               'infak_terikat':'Infak Terikat','amil':'Amil','wakaf':'Wakaf','umum':'Umum'}
LABEL_ASNAF = {'fakir':'Fakir','miskin':'Miskin','amil':'Amil','muallaf':'Muallaf',
               'riqab':'Riqab','gharim':'Gharim','fisabilillah':'Fisabilillah','ibnu_sabil':'Ibnu Sabil'}
LABEL_SUMBER = {'tunai':'Tunai','kencleng':'Kencleng','kotak_infaq':'Kotak Infaq'}
BULAN_IND   = {1:'Januari',2:'Februari',3:'Maret',4:'April',5:'Mei',6:'Juni',
               7:'Juli',8:'Agustus',9:'September',10:'Oktober',11:'November',12:'Desember'}

def format_bulan(b):
    try:
        y, m = b.split('-')
        return f"{BULAN_IND[int(m)]} {y}"
    except:
        return b

app.jinja_env.filters['bulan_label'] = format_bulan
app.jinja_env.globals.update(LABEL_DANA=LABEL_DANA, LABEL_ASNAF=LABEL_ASNAF,
                              LABEL_SUMBER=LABEL_SUMBER, DANA_TYPES=DANA_TYPES,
                              KODE_KAS_DANA=KODE_KAS_DANA, KODE_SALDO_DANA=KODE_SALDO_DANA,
                              BULAN_IND=BULAN_IND)

def parse_gmaps_url(url):
    """Ekstrak (lat, lng) dari berbagai format URL Google Maps."""
    patterns = [
        r'@(-?\d+\.\d+),(-?\d+\.\d+)',
        r'[?&]q=(-?\d+\.\d+),(-?\d+\.\d+)',
        r'[?&]ll=(-?\d+\.\d+),(-?\d+\.\d+)',
    ]
    for p in patterns:
        m = re.search(p, url)
        if m:
            return float(m.group(1)), float(m.group(2))
    return None, None

def buka_periode(bulan, user_id):
    """Generate koleksi_bulanan untuk semua donatur aktif bulan tsb."""
    conn = get_db()
    donatur = conn.execute(
        "SELECT id FROM donatur WHERE sumber_infaq IN ('kencleng','kotak_infaq') "
        "AND aktif_infaq=1 AND aktif=1"
    ).fetchall()
    created = 0
    for d in donatur:
        try:
            conn.execute("INSERT INTO koleksi_bulanan (donatur_id, bulan) VALUES (?,?)",
                         (d['id'], bulan))
            created += 1
        except sqlite3.IntegrityError:
            pass
    conn.commit()
    conn.close()
    return created

def auto_koleksi_donatur_baru(conn, donatur_id, sumber_infaq):
    """Buat koleksi_bulanan bulan ini untuk donatur baru kencleng/kotak_infaq."""
    if sumber_infaq in ('kencleng', 'kotak_infaq'):
        bulan = date.today().strftime('%Y-%m')
        try:
            conn.execute("INSERT INTO koleksi_bulanan (donatur_id, bulan) VALUES (?,?)",
                         (donatur_id, bulan))
        except:
            pass

def get_instansi(conn=None):
    close = False
    if conn is None:
        conn = get_db(); close = True
    row = conn.execute("SELECT * FROM instansi WHERE id=1").fetchone()
    if close: conn.close()
    if row:
        return dict(row)
    return {'nama': 'BAITUL MAAL BMT', 'nama_lembaga': '', 'alamat': '', 'telepon': '',
            'email': '', 'website': '', 'ketua': '', 'bendahara': '', 'sekretaris': '',
            'no_izin': ''}

def get_saldo_awal(conn):
    """Saldo awal manual per jenis dana (sebelum transaksi pertama tercatat di sistem)."""
    rows = conn.execute("SELECT jenis_dana, jumlah FROM saldo_awal").fetchall()
    return {r['jenis_dana']: r['jumlah'] for r in rows}

def auto_transaksi_koleksi(conn, koleksi_id, donatur_id, bulan, sumber, jumlah, tanggal, user_id):
    """Buat transaksi otomatis saat koleksi terkumpul."""
    coa = conn.execute("SELECT id, jenis_dana FROM chart_of_accounts WHERE kode='4.2.2'").fetchone()
    keterangan = f"Koleksi {LABEL_SUMBER.get(sumber, sumber)} – {bulan}"
    cur = conn.execute("""
        INSERT INTO transaksi (tanggal,jenis,jenis_dana,coa_id,donatur_id,jumlah,keterangan,user_id)
        VALUES (?,?,?,?,?,?,?,?)
    """, (tanggal, 'masuk', coa['jenis_dana'] if coa else None, coa['id'] if coa else None,
          donatur_id, jumlah, keterangan, user_id))
    trx_id = cur.lastrowid
    conn.execute("UPDATE koleksi_bulanan SET transaksi_id=? WHERE id=?", (trx_id, koleksi_id))
    return trx_id

# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route('/', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('admin_dashboard') if session['role']=='admin' else url_for('marketing_dashboard'))
    error = None
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        conn = get_db()
        user = conn.execute(
            "SELECT * FROM users WHERE username=? AND password=? AND aktif=1",
            (username, hash_pw(password))
        ).fetchone()
        conn.close()
        if user:
            session.update({'user_id':user['id'],'username':user['username'],
                            'nama':user['nama'],'role':user['role']})
            return redirect(url_for('admin_dashboard') if user['role']=='admin' else url_for('marketing_dashboard'))
        error = 'Username atau password salah.'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── Admin Dashboard ───────────────────────────────────────────────────────────

@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = get_db()
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    total_masuk  = conn.execute("SELECT COALESCE(SUM(jumlah),0) FROM transaksi WHERE jenis='masuk' AND strftime('%Y-%m',tanggal)=?", (bulan,)).fetchone()[0]
    total_keluar = conn.execute("SELECT COALESCE(SUM(jumlah),0) FROM transaksi WHERE jenis='keluar' AND strftime('%Y-%m',tanggal)=?", (bulan,)).fetchone()[0]
    saldo = conn.execute("SELECT COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE -jumlah END),0) FROM transaksi").fetchone()[0]
    transaksi_terakhir = conn.execute('''
        SELECT t.*, c.nama as coa_nama, c.jenis_dana, d.nama as donatur_nama, u.nama as petugas
        FROM transaksi t
        LEFT JOIN chart_of_accounts c ON t.coa_id=c.id
        LEFT JOIN donatur d ON t.donatur_id=d.id
        LEFT JOIN users u ON t.user_id=u.id
        ORDER BY t.created_at DESC LIMIT 10
    ''').fetchall()
    rekap_dana = conn.execute('''
        SELECT jenis_dana, jenis, SUM(jumlah) as total FROM transaksi
        WHERE strftime('%Y-%m',tanggal)=? AND jenis_dana IS NOT NULL
        GROUP BY jenis_dana, jenis ORDER BY jenis_dana
    ''', (bulan,)).fetchall()
    conn.close()
    return render_template('admin/dashboard.html', total_masuk=total_masuk,
        total_keluar=total_keluar, saldo=saldo, transaksi_terakhir=transaksi_terakhir,
        rekap_dana=rekap_dana, bulan=bulan)

# ── Admin Transaksi ───────────────────────────────────────────────────────────

MUSTAHIK_RE = re.compile(r'\((\d+)\s*mustahik\)', re.IGNORECASE)

@app.route('/admin/transaksi')
@admin_required
def admin_transaksi():
    conn = get_db()
    bulan = request.args.get('bulan', get_tanggal_kerja()[:7])
    jenis = request.args.get('jenis', '')
    jenis_dana = request.args.get('jenis_dana', '')
    perlu_lengkap = request.args.get('perlu_lengkap', '')
    query = '''
        SELECT t.*, c.nama as coa_nama, c.kode as coa_kode, c.jenis_dana,
               d.nama as donatur_nama, p.nama as penerima_nama, u.nama as petugas
        FROM transaksi t
        LEFT JOIN chart_of_accounts c ON t.coa_id=c.id
        LEFT JOIN donatur d ON t.donatur_id=d.id
        LEFT JOIN penerima_manfaat p ON t.penerima_id=p.id
        LEFT JOIN users u ON t.user_id=u.id
        WHERE strftime('%Y-%m',t.tanggal)=?'''
    params = [bulan]
    if jenis in ('masuk','keluar'):
        query += ' AND t.jenis=?'; params.append(jenis)
    if jenis_dana:
        query += ' AND t.jenis_dana=?'; params.append(jenis_dana)
    if perlu_lengkap:
        query += " AND t.jenis='keluar' AND t.jumlah_mustahik IS NULL AND t.jurnal_id IS NULL"
    query += ' ORDER BY t.tanggal DESC, t.created_at DESC'
    transaksi   = conn.execute(query, params).fetchall()
    tebakan_mustahik = {}
    if perlu_lengkap:
        for t in transaksi:
            m = MUSTAHIK_RE.search(t['keterangan'] or '')
            if m:
                tebakan_mustahik[t['id']] = int(m.group(1))
    coa_list    = conn.execute("SELECT * FROM chart_of_accounts WHERE jenis_transaksi IS NOT NULL AND aktif=1 ORDER BY kode").fetchall()
    coa_parents = conn.execute(
        "SELECT * FROM chart_of_accounts WHERE parent_kode IS NOT NULL AND aktif=1 ORDER BY kode"
    ).fetchall()
    donatur_list= conn.execute("SELECT * FROM donatur WHERE aktif=1 ORDER BY nama").fetchall()
    penerima_list=conn.execute("SELECT * FROM penerima_manfaat WHERE aktif=1 ORDER BY nama").fetchall()
    conn.close()
    return render_template('admin/transaksi.html', transaksi=transaksi, coa_list=coa_list,
        coa_parents=coa_parents, donatur_list=donatur_list, penerima_list=penerima_list,
        bulan=bulan, jenis=jenis, jenis_dana=jenis_dana, perlu_lengkap=perlu_lengkap,
        tebakan_mustahik=tebakan_mustahik)

@app.route('/admin/transaksi/<int:id>/mustahik', methods=['POST'])
@admin_required
def set_jumlah_mustahik(id):
    """Isi/perbaiki jumlah_mustahik pada transaksi keluar yang sudah ada — dipakai utk
    backfill data lama yg dicatat sebelum field ini ada. Sengaja dibatasi hanya kolom ini
    (bukan edit transaksi umum) krn belum ada fitur edit transaksi sama sekali di sistem."""
    conn = get_db()
    row = conn.execute("SELECT jenis FROM transaksi WHERE id=?", (id,)).fetchone()
    if not row or row['jenis'] != 'keluar':
        conn.close()
        flash('Transaksi tidak ditemukan atau bukan penyaluran.', 'danger')
        return redirect(url_for('admin_transaksi'))
    raw = request.form.get('jumlah_mustahik', '').strip()
    nilai = None
    if raw:
        try:
            nilai = max(0, int(raw))
        except ValueError:
            flash('Jumlah mustahik harus angka.', 'danger')
            conn.close()
            return redirect(request.referrer or url_for('admin_transaksi'))
    conn.execute("UPDATE transaksi SET jumlah_mustahik=? WHERE id=?", (nilai, id))
    conn.commit(); conn.close()
    flash('Jumlah mustahik disimpan.', 'success')
    return redirect(request.referrer or url_for('admin_transaksi'))

@app.route('/admin/transaksi/tambah', methods=['POST'])
@admin_required
def tambah_transaksi():
    data = request.form
    if data.get('jenis') not in ('masuk', 'keluar'):
        flash('Pilih Sub Akun terlebih dahulu (jenis transaksi belum terisi).', 'danger')
        return redirect(url_for('admin_transaksi'))
    jumlah = parse_jumlah(data.get('jumlah'))
    if jumlah is None:
        flash('Jumlah tidak valid — harus angka lebih dari 0.', 'danger')
        return redirect(url_for('admin_transaksi'))
    jumlah_mustahik = None
    if data.get('jumlah_mustahik'):
        try: jumlah_mustahik = int(data['jumlah_mustahik'])
        except ValueError: jumlah_mustahik = None
    conn = get_db()
    trx_id, dup = insert_transaksi(conn, data['tanggal'], data['jenis'],
        data.get('coa_id') or None, data.get('donatur_id') or None,
        data.get('penerima_id') or None, jumlah,
        data.get('keterangan',''), session['user_id'], data.get('client_uuid'),
        nama_kegiatan=data.get('nama_kegiatan') or None,
        lokasi=data.get('lokasi') or None,
        jumlah_mustahik=jumlah_mustahik)
    conn.commit(); conn.close()
    if dup:
        flash('Transaksi ini sudah tercatat sebelumnya — tidak dicatat ganda.', 'warning')
    else:
        flash('Transaksi berhasil dicatat.', 'success')
    return redirect(url_for('admin_transaksi'))

@app.route('/admin/transaksi/hapus/<int:id>', methods=['POST'])
@admin_required
def hapus_transaksi(id):
    conn = get_db()
    conn.execute("DELETE FROM transaksi WHERE id=?", (id,))
    conn.commit(); conn.close()
    flash('Transaksi dihapus.', 'warning')
    return redirect(url_for('admin_transaksi'))

# ── Admin Laporan ─────────────────────────────────────────────────────────────

@app.route('/admin/laporan')
@admin_required
def admin_laporan():
    return render_template('admin/laporan.html')


def _last_day(bulan):
    y, m = map(int, bulan.split('-'))
    return f"{bulan}-{cal_mod.monthrange(y, m)[1]:02d}"

def _prev_last_day(bulan):
    y, m = map(int, bulan.split('-'))
    if m == 1: py, pm = y-1, 12
    else:      py, pm = y,   m-1
    return f"{py}-{pm:02d}-{cal_mod.monthrange(py, pm)[1]:02d}"

def _dana_summary(conn, bulan):
    """Hitung saldo awal, penerimaan, penyaluran, saldo akhir per jenis dana."""
    dana_types = DANA_TYPES
    saldo_manual = get_saldo_awal(conn)

    rows_awal = conn.execute("""
        SELECT jenis_dana,
               COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE -jumlah END),0) as saldo
        FROM transaksi WHERE strftime('%Y-%m',tanggal) < ? AND jenis_dana IS NOT NULL
        GROUP BY jenis_dana
    """, (bulan,)).fetchall()
    saldo_awal = {r['jenis_dana']: saldo_manual.get(r['jenis_dana'], 0) + r['saldo'] for r in rows_awal}
    for dana in dana_types:
        saldo_awal.setdefault(dana, saldo_manual.get(dana, 0))

    masuk_rows = conn.execute("""
        SELECT c.kode, c.nama, c.jenis_dana, c.parent_kode, SUM(t.jumlah) as total
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id=c.id
        WHERE t.jenis='masuk' AND strftime('%Y-%m',t.tanggal)=?
        GROUP BY c.id ORDER BY c.kode
    """, (bulan,)).fetchall()

    keluar_rows = conn.execute("""
        SELECT c.kode, c.nama, c.jenis_dana, c.parent_kode, SUM(t.jumlah) as total
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id=c.id
        WHERE t.jenis='keluar' AND strftime('%Y-%m',t.tanggal)=?
        GROUP BY c.id ORDER BY c.kode
    """, (bulan,)).fetchall()

    data = {}
    for dana in dana_types:
        masuk  = [r for r in masuk_rows  if r['jenis_dana'] == dana]
        keluar = [r for r in keluar_rows if r['jenis_dana'] == dana]
        tm = sum(r['total'] for r in masuk)
        tk = sum(r['total'] for r in keluar)
        sa = saldo_awal.get(dana, 0)
        data[dana] = {
            'masuk': masuk, 'keluar': keluar,
            'total_masuk': tm, 'total_keluar': tk,
            'saldo_awal': sa, 'saldo_akhir': sa + tm - tk,
        }
    return data


@app.route('/admin/laporan/neraca')
@admin_required
def laporan_neraca():
    bulan = request.args.get('bulan', get_tanggal_kerja()[:7])
    ld    = _last_day(bulan)
    conn  = get_db()

    rows = conn.execute("""
        SELECT jenis_dana,
               COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE -jumlah END),0) as saldo
        FROM transaksi WHERE tanggal <= ? AND jenis_dana IS NOT NULL
        GROUP BY jenis_dana
    """, (ld,)).fetchall()
    kas_trans = {r['jenis_dana']: r['saldo'] for r in rows}
    saldo_manual = get_saldo_awal(conn)

    dana_types = DANA_TYPES
    kas = {d: kas_trans.get(d, 0) + saldo_manual.get(d, 0) for d in dana_types}
    total_aset = sum(kas.get(d, 0) for d in dana_types)
    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_neraca.html',
        kas=kas, dana_types=dana_types, total_aset=total_aset,
        bulan=bulan, last_day=ld, inst=inst)


@app.route('/admin/laporan/dana')
@admin_required
def laporan_dana():
    bulan = request.args.get('bulan', get_tanggal_kerja()[:7])
    conn  = get_db()
    data  = _dana_summary(conn, bulan)
    inst  = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_dana.html',
        data=data, bulan=bulan, inst=inst, dana_types=DANA_TYPES)


def _program_code(nama):
    """Ambil kode program dari akhiran '[KODE]' pada nama akun, mis. 'Cinta Yatim [CY]' -> 'CY'."""
    m = re.search(r"\[([A-Za-z0-9']+)\]\s*$", nama or '')
    return m.group(1).upper() if m else None


def _hitung_saldo_program(conn, bulan):
    """Hitung saldo per program (gabungan sisi penerimaan+penyaluran akun berkode sama,
    mis. [CY]) s/d akhir `bulan`. Dipakai bareng oleh laporan Saldo per Program & API
    saldo akun (popup di menu Tambah Transaksi) spy logikanya selalu konsisten.
    Return (groups, tidak_terikat, coa_key):
      - groups: {key: {..., saldo_akhir}} per program/akun individual (SEBELUM akun
        tidak-terikat digabung & SEBELUM key 'Miskin' dibuang oleh alokasi zakat)
      - tidak_terikat: dict agregat 1 baris utk semua akun berjenis infak_tidak_terikat
      - coa_key: {coa_id: key} utk lookup akun -> key aslinya di `groups`
    """
    fd = f"{bulan}-01"
    ld = _last_day(bulan)

    saldo_awal_rows = conn.execute("""
        SELECT c.id as coa_id, c.nama, c.jenis_dana, c.kelompok,
               COALESCE(SUM(CASE WHEN t.jenis='masuk' THEN t.jumlah ELSE -t.jumlah END),0) as saldo
        FROM chart_of_accounts c
        LEFT JOIN transaksi t ON t.coa_id=c.id AND t.tanggal < ?
        WHERE c.jenis_transaksi IS NOT NULL AND c.aktif=1
        GROUP BY c.id
    """, (fd,)).fetchall()

    period_rows = conn.execute("""
        SELECT c.id as coa_id,
               COALESCE(SUM(CASE WHEN t.jenis='masuk' THEN t.jumlah ELSE 0 END),0) as masuk,
               COALESCE(SUM(CASE WHEN t.jenis='keluar' THEN t.jumlah ELSE 0 END),0) as keluar
        FROM chart_of_accounts c
        LEFT JOIN transaksi t ON t.coa_id=c.id AND t.tanggal BETWEEN ? AND ?
        WHERE c.jenis_transaksi IS NOT NULL AND c.aktif=1
        GROUP BY c.id
    """, (fd, ld)).fetchall()
    period_map = {r['coa_id']: r for r in period_rows}

    groups = {}
    coa_key = {}
    for r in saldo_awal_rows:
        code = _program_code(r['nama'])
        key = code or f"coa{r['coa_id']}"
        coa_key[r['coa_id']] = key
        if key not in groups:
            label = re.sub(r"\s*\[.*?\]\s*$", '', r['nama']).strip()
            groups[key] = {'label': label, 'code': code, 'jenis_dana': r['jenis_dana'],
                            'saldo_awal': 0, 'masuk': 0, 'keluar': 0}
        g = groups[key]
        g['saldo_awal'] += r['saldo']
        p = period_map.get(r['coa_id'])
        if p:
            g['masuk']  += p['masuk']
            g['keluar'] += p['keluar']
        # program gabungan bisa punya jenis_dana beda di sisi penerimaan vs penyaluran
        # (mis. akun penyaluran lama masih tertandai tidak-terikat) -> pakai jenis_dana sisi penerimaan
        if r['kelompok'] == 'penerimaan':
            g['jenis_dana'] = r['jenis_dana']

    # Alokasi penyaluran "Miskin" (dr Zakat): habiskan dulu saldo Fidyah, lalu Zakat
    # Fitrah, sisanya dr Zakat Maal — bukan transaksi baru, murni cara tampil krn
    # ketiganya sama2 dana zakat & Miskin tdk py sumber spesifik per transaksi.
    zakat_kode = {r['kode']: r['id'] for r in conn.execute(
        "SELECT kode, id FROM chart_of_accounts WHERE kode IN ('4.1.1','4.1.2','4.1.6','5.1.2')"
    ).fetchall()}
    fid_key = 'F'
    zf_key  = f"coa{zakat_kode.get('4.1.2')}"
    zm_key  = f"coa{zakat_kode.get('4.1.1')}"
    miskin_key = f"coa{zakat_kode.get('5.1.2')}"
    if miskin_key in groups:
        miskin_g = groups.pop(miskin_key)
        sisa = -(miskin_g['saldo_awal'] + miskin_g['masuk'] - miskin_g['keluar'])
        for k in (fid_key, zf_key, zm_key):
            if sisa <= 0 or k not in groups:
                continue
            g = groups[k]
            saldo_sblm_alokasi = g['saldo_awal'] + g['masuk'] - g['keluar']
            potong = min(sisa, saldo_sblm_alokasi) if saldo_sblm_alokasi > 0 else 0
            g['keluar'] += potong
            sisa -= potong
        if sisa > 0 and zm_key in groups:
            groups[zm_key]['keluar'] += sisa  # kekurangan (jika ada) tetap nempel Zakat Maal

    for g in groups.values():
        g['saldo_akhir'] = g['saldo_awal'] + g['masuk'] - g['keluar']

    # Infak Tidak Terikat dikelola sbg satu saldo gabungan (Kotak Infaq+Kencleng+Tunai
    # di penerimaan; Kafalah Guru TPQ, Safari Masjid, Sembako Dhuafa, Hibah ke Dana
    # Lain, dll di penyaluran) — bukan program per-akun, jadi digabung jadi 1 baris.
    tidak_terikat = {'label': 'Infaq Tidak Terikat (gabungan)', 'code': None,
                      'jenis_dana': 'infak_tidak_terikat', 'saldo_awal': 0, 'masuk': 0, 'keluar': 0}
    for g in groups.values():
        if g['jenis_dana'] == 'infak_tidak_terikat':
            tidak_terikat['saldo_awal'] += g['saldo_awal']
            tidak_terikat['masuk']      += g['masuk']
            tidak_terikat['keluar']     += g['keluar']
    tidak_terikat['saldo_akhir'] = tidak_terikat['saldo_awal'] + tidak_terikat['masuk'] - tidak_terikat['keluar']

    return groups, tidak_terikat, coa_key


def _saldo_program_list(groups, tidak_terikat):
    """Ratakan output _hitung_saldo_program() jadi list siap-tampil: akun2 tidak
    terikat digabung jadi 1 baris, akun2 Amil (penerimaan bagian amil + beban gaji/
    operasional/dll) jg digabung jadi 1 baris -- itu kategori beban umum, bukan
    program tersendiri, jd gak perlu tampil kepotong2 pemasukan/pengeluaran sendiri2.
    Urut per jenis dana lalu saldo akhir terbesar."""
    data = []
    amil = {'label': 'Dana Amil (gabungan)', 'code': None,
            'jenis_dana': 'amil', 'saldo_awal': 0, 'masuk': 0, 'keluar': 0}
    for g in groups.values():
        if g['jenis_dana'] == 'infak_tidak_terikat':
            continue
        if g['jenis_dana'] == 'amil':
            amil['saldo_awal'] += g['saldo_awal']
            amil['masuk']      += g['masuk']
            amil['keluar']     += g['keluar']
            continue
        if g['saldo_awal'] or g['masuk'] or g['keluar']:
            data.append(g)
    amil['saldo_akhir'] = amil['saldo_awal'] + amil['masuk'] - amil['keluar']
    if tidak_terikat['saldo_awal'] or tidak_terikat['masuk'] or tidak_terikat['keluar']:
        data.append(tidak_terikat)
    if amil['saldo_awal'] or amil['masuk'] or amil['keluar']:
        data.append(amil)
    data.sort(key=lambda g: (DANA_TYPES.index(g['jenis_dana']) if g['jenis_dana'] in DANA_TYPES else 99, -g['saldo_akhir']))
    return data


@app.route('/admin/laporan/saldo-program')
@admin_required
def laporan_saldo_program():
    """Saldo per program/produk (gabungan sisi penerimaan+penyaluran akun berkode sama, mis. [CY])."""
    bulan = request.args.get('bulan', get_tanggal_kerja()[:7])
    conn = get_db()
    groups, tidak_terikat, _ = _hitung_saldo_program(conn, bulan)
    data = _saldo_program_list(groups, tidak_terikat)

    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_saldo_program.html',
        data=data, bulan=bulan, inst=inst, dana_types=DANA_TYPES)


@app.route('/api/coa/<int:coa_id>/saldo')
@admin_required
def api_coa_saldo(coa_id):
    """Sisa saldo program dari akun terpilih, utk popup info di menu Tambah Transaksi."""
    bulan = request.args.get('bulan', get_tanggal_kerja()[:7])
    conn = get_db()
    akun = conn.execute(
        "SELECT id, kode, nama, jenis_dana FROM chart_of_accounts WHERE id=?", (coa_id,)
    ).fetchone()
    if not akun:
        conn.close()
        return jsonify(ok=False), 404

    groups, tidak_terikat, coa_key = _hitung_saldo_program(conn, bulan)
    conn.close()

    if akun['jenis_dana'] == 'infak_tidak_terikat':
        return jsonify(ok=True, label='Infaq Tidak Terikat (gabungan)',
                        saldo=tidak_terikat['saldo_akhir'],
                        catatan='dikelola sbg satu saldo gabungan (Kotak Infaq+Kencleng+Tunai)')

    if akun['kode'] == '5.1.2':  # Miskin: dialokasikan otomatis dr Fidyah/ZF/ZM, tdk py saldo sendiri
        total_zakat = sum(g['saldo_akhir'] for g in groups.values() if g['jenis_dana'] == 'zakat')
        return jsonify(ok=True, label='Total Dana Zakat (Fidyah + Zakat Fitrah + Zakat Maal)',
                        saldo=total_zakat,
                        catatan='Miskin dialokasikan otomatis dr saldo zakat, tdk py saldo terpisah')

    key = coa_key.get(coa_id)
    g = groups.get(key)
    if not g:
        return jsonify(ok=False)
    label = g['label'] + (f" [{g['code']}]" if g['code'] else '')
    return jsonify(ok=True, label=label, saldo=g['saldo_akhir'])


def _program_registry(conn):
    """Daftar semua program (gabungan akun berkode sama, mis. [CY]) + daftar coa_id
    anggotanya -- dipakai Buku Besar Program utk narik transaksi lintas-akun 1 program.
    Semua akun infak_tidak_terikat digabung jadi satu key 'TIDAK_TERIKAT' (dikelola
    1 saldo, sama seperti di Saldo per Program)."""
    rows = conn.execute("""
        SELECT id, kode, nama, jenis_dana, kelompok FROM chart_of_accounts
        WHERE jenis_transaksi IS NOT NULL AND aktif=1
    """).fetchall()
    programs = {}
    for r in rows:
        code = _program_code(r['nama'])
        key = code or f"coa{r['id']}"
        if key not in programs:
            label = re.sub(r"\s*\[.*?\]\s*$", '', r['nama']).strip()
            programs[key] = {'label': label, 'code': code, 'jenis_dana': r['jenis_dana'], 'coa_ids': []}
        p = programs[key]
        p['coa_ids'].append(r['id'])
        if r['kelompok'] == 'penerimaan':
            p['jenis_dana'] = r['jenis_dana']

    tt_ids = []
    for key in [k for k, p in programs.items() if p['jenis_dana'] == 'infak_tidak_terikat']:
        tt_ids += programs.pop(key)['coa_ids']
    if tt_ids:
        programs['TIDAK_TERIKAT'] = {'label': 'Infaq Tidak Terikat (gabungan)', 'code': None,
                                      'jenis_dana': 'infak_tidak_terikat', 'coa_ids': tt_ids}
    return programs


@app.route('/admin/laporan/buku-besar')
@admin_required
def laporan_buku_besar():
    """Buku besar per program: rincian mutasi + saldo berjalan, utk semua program
    (Zakat, Infaq Tidak Terikat gabungan, maupun tiap program Infaq Terikat).
    Rentang tanggal bebas (bkn cuma 1 bulan) spy bisa akses dr awal tahun."""
    tk = get_tanggal_kerja()
    fd = request.args.get('dari', f"{tk[:4]}-01-01")
    ld = request.args.get('sampai', tk)
    program_key = request.args.get('program', '')
    conn = get_db()
    programs = _program_registry(conn)

    ringkasan = []
    for key, p in programs.items():
        ph = ','.join('?' * len(p['coa_ids']))
        saldo_awal = conn.execute(
            f"SELECT COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE -jumlah END),0) "
            f"FROM transaksi WHERE coa_id IN ({ph}) AND tanggal < ?",
            (*p['coa_ids'], fd)
        ).fetchone()[0]
        agg = conn.execute(
            f"SELECT COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE 0 END),0) as masuk, "
            f"COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END),0) as keluar "
            f"FROM transaksi WHERE coa_id IN ({ph}) AND tanggal BETWEEN ? AND ?",
            (*p['coa_ids'], fd, ld)
        ).fetchone()
        saldo_akhir = saldo_awal + agg['masuk'] - agg['keluar']
        if saldo_awal or agg['masuk'] or agg['keluar']:
            ringkasan.append({'key': key, 'label': p['label'], 'code': p['code'], 'jenis_dana': p['jenis_dana'],
                               'saldo_awal': saldo_awal, 'masuk': agg['masuk'], 'keluar': agg['keluar'],
                               'saldo_akhir': saldo_akhir})
    ringkasan.sort(key=lambda g: (DANA_TYPES.index(g['jenis_dana']) if g['jenis_dana'] in DANA_TYPES else 99, -g['saldo_akhir']))

    subtotal_per_dana = {}
    for g in ringkasan:
        s = subtotal_per_dana.setdefault(g['jenis_dana'], {'saldo_awal': 0, 'masuk': 0, 'keluar': 0, 'saldo_akhir': 0})
        s['saldo_awal']  += g['saldo_awal']
        s['masuk']       += g['masuk']
        s['keluar']      += g['keluar']
        s['saldo_akhir'] += g['saldo_akhir']

    detail = None
    if program_key and program_key in programs:
        p = programs[program_key]
        ph = ','.join('?' * len(p['coa_ids']))
        saldo_awal = conn.execute(
            f"SELECT COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE -jumlah END),0) "
            f"FROM transaksi WHERE coa_id IN ({ph}) AND tanggal < ?",
            (*p['coa_ids'], fd)
        ).fetchone()[0]
        rows = conn.execute(
            f"SELECT t.id, t.tanggal, t.jenis, t.jumlah, t.keterangan, c.kode as coa_kode, c.nama as coa_nama, "
            f"d.nama as donatur_nama, pm.nama as penerima_nama "
            f"FROM transaksi t JOIN chart_of_accounts c ON t.coa_id=c.id "
            f"LEFT JOIN donatur d ON t.donatur_id=d.id "
            f"LEFT JOIN penerima_manfaat pm ON t.penerima_id=pm.id "
            f"WHERE t.coa_id IN ({ph}) AND t.tanggal BETWEEN ? AND ? "
            f"ORDER BY t.tanggal, t.id",
            (*p['coa_ids'], fd, ld)
        ).fetchall()
        baris = []
        running = saldo_awal
        total_masuk = total_keluar = 0
        for r in rows:
            if r['jenis'] == 'masuk':
                running += r['jumlah']; total_masuk += r['jumlah']
            else:
                running -= r['jumlah']; total_keluar += r['jumlah']
            baris.append({**dict(r), 'saldo': running})
        detail = {'key': program_key, 'label': p['label'], 'code': p['code'], 'jenis_dana': p['jenis_dana'],
                   'saldo_awal': saldo_awal, 'baris': baris, 'total_masuk': total_masuk,
                   'total_keluar': total_keluar, 'saldo_akhir': running}

    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_buku_besar.html',
        ringkasan=ringkasan, detail=detail, program_key=program_key, subtotal_per_dana=subtotal_per_dana,
        dari=fd, sampai=ld, awal_tahun=f"{tk[:4]}-01-01", awal_bulan=f"{tk[:7]}-01", hari_ini=tk,
        inst=inst, dana_types=DANA_TYPES)


def _rekap_penerimaan_tahunan(conn, parent_kode, tahun):
    """Rekap tahunan penerimaan per akun anak dari `parent_kode`, per bulan (baris) x
    per akun (kolom) -- bahan rapat. Dipakai bareng oleh rekap Infaq Tidak Terikat,
    Zakat, & Infaq Terikat spy formatnya konsisten."""
    sumber_coa = conn.execute("""
        SELECT id, kode, nama FROM chart_of_accounts
        WHERE parent_kode=? AND jenis_transaksi='masuk' AND aktif=1 ORDER BY kode
    """, (parent_kode,)).fetchall()

    rows = conn.execute("""
        SELECT strftime('%m', t.tanggal) as bln, t.coa_id, SUM(t.jumlah) as total
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id=c.id
        WHERE c.parent_kode=? AND t.jenis='masuk' AND strftime('%Y', t.tanggal)=?
        GROUP BY bln, t.coa_id
    """, (parent_kode, tahun)).fetchall()
    per_bulan = {}
    for r in rows:
        per_bulan.setdefault(r['bln'], {})[r['coa_id']] = r['total']

    bulan_urut = [f"{i:02d}" for i in range(1, 13)]
    tabel = []
    total_per_sumber = {c['id']: 0 for c in sumber_coa}
    grand_total = 0
    for bm in bulan_urut:
        vals = per_bulan.get(bm, {})
        if not vals:
            continue
        baris = {'bulan': bm, 'label': BULAN_IND[int(bm)], 'per_sumber': {}, 'total': 0}
        for c in sumber_coa:
            v = vals.get(c['id'], 0)
            baris['per_sumber'][c['id']] = v
            baris['total'] += v
            total_per_sumber[c['id']] += v
        grand_total += baris['total']
        tabel.append(baris)

    return sumber_coa, tabel, total_per_sumber, grand_total


@app.route('/admin/laporan/rekap-sumber-infaq')
@admin_required
def laporan_rekap_sumber_infaq():
    """Rekap tahunan penerimaan Infak Tidak Terikat per sumber (Kotak Infaq/Kencleng/Tunai) —
    bahan rapat. Sesuai pengelolaan riil: ketiganya jadi satu saldo, ini murni rekap sumber dana."""
    tahun = request.args.get('tahun', get_tanggal_kerja()[:4])
    conn = get_db()
    sumber_coa, tabel, total_per_sumber, grand_total = _rekap_penerimaan_tahunan(conn, '4.2.2', tahun)

    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_rekap_sumber.html',
        sumber_coa=sumber_coa, tabel=tabel, total_per_sumber=total_per_sumber,
        grand_total=grand_total, tahun=tahun, inst=inst,
        judul='PENERIMAAN INFAQ BEBAS PER SUMBER',
        deskripsi='Rincian sumber penerimaan Infaq Bebas/Tidak Terikat (Kotak Infaq, Kencleng, Tunai) '
                  'untuk bahan rapat — dalam pengelolaan sehari-hari ketiganya tetap satu saldo.',
    )


@app.route('/admin/laporan/rekap-zakat')
@admin_required
def laporan_rekap_zakat():
    """Rekap tahunan penerimaan Zakat per jenis (Zakat Maal, Fitrah, Fidyah, dst) — bahan rapat."""
    tahun = request.args.get('tahun', get_tanggal_kerja()[:4])
    conn = get_db()
    sumber_coa, tabel, total_per_sumber, grand_total = _rekap_penerimaan_tahunan(conn, '4.1', tahun)
    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_rekap_sumber.html',
        sumber_coa=sumber_coa, tabel=tabel, total_per_sumber=total_per_sumber,
        grand_total=grand_total, tahun=tahun, inst=inst,
        judul='PENERIMAAN ZAKAT PER JENIS',
        deskripsi='Rincian penerimaan Zakat per jenis (Zakat Maal, Zakat Fitrah, Zakat Penghasilan/Profesi, '
                  'Fidyah, dst) per bulan selama setahun — bahan rapat.',
    )


@app.route('/admin/laporan/rekap-infaq-terikat')
@admin_required
def laporan_rekap_infaq_terikat():
    """Rekap tahunan penerimaan Infaq Terikat per program (Yatim, Ambulance, Qurban, dst) — bahan rapat."""
    tahun = request.args.get('tahun', get_tanggal_kerja()[:4])
    conn = get_db()
    sumber_coa, tabel, total_per_sumber, grand_total = _rekap_penerimaan_tahunan(conn, '4.2.1', tahun)
    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_rekap_sumber.html',
        sumber_coa=sumber_coa, tabel=tabel, total_per_sumber=total_per_sumber,
        grand_total=grand_total, tahun=tahun, inst=inst,
        judul='PENERIMAAN INFAQ TERIKAT PER PROGRAM',
        deskripsi='Rincian penerimaan tiap program Infaq Terikat (Yatim, Ambulance, Sunat Sehat Gratis, '
                  'Qurban, dst) per bulan selama setahun — bahan rapat.',
    )


def _rekap_donasi_donatur(conn, tahun):
    """Rekap tahunan penerimaan per donatur (baris) x per bulan (kolom) + statistik --
    bahan evaluasi akhir tahun. Hanya donatur dgn min. 1 transaksi masuk di tahun tsb
    yg ditampilkan (dari 1000+ donatur, kebanyakan tak aktif tiap bulan)."""
    rows = conn.execute("""
        SELECT d.id, d.nama, d.sumber_infaq, d.aktif,
               strftime('%m', t.tanggal) as bln, SUM(t.jumlah) as total
        FROM transaksi t JOIN donatur d ON t.donatur_id = d.id
        WHERE t.jenis='masuk' AND strftime('%Y', t.tanggal)=?
        GROUP BY d.id, bln
    """, (tahun,)).fetchall()

    donatur_map = {}
    for r in rows:
        e = donatur_map.setdefault(r['id'], {
            'id': r['id'], 'nama': r['nama'], 'sumber_infaq': r['sumber_infaq'],
            'aktif': r['aktif'], 'per_bulan': {}, 'total': 0, 'bulan_aktif': 0,
        })
        e['per_bulan'][r['bln']] = r['total']
        e['total'] += r['total']
        e['bulan_aktif'] += 1

    tabel = list(donatur_map.values())
    for e in tabel:
        e['rata2'] = e['total'] / e['bulan_aktif'] if e['bulan_aktif'] else 0
    tabel.sort(key=lambda e: -e['total'])

    total_per_bulan = {f"{i:02d}": 0 for i in range(1, 13)}
    grand_total = 0
    for e in tabel:
        for bm, v in e['per_bulan'].items():
            total_per_bulan[bm] += v
        grand_total += e['total']

    return tabel, total_per_bulan, grand_total


@app.route('/admin/laporan/rekap-donasi-donatur')
@admin_required
def laporan_rekap_donasi_donatur():
    """Rekap tahunan penerimaan per donatur x bulan + statistik (bln aktif, rata-rata) --
    bahan evaluasi akhir tahun (donatur konsisten vs jarang/berhenti)."""
    tahun = request.args.get('tahun', get_tanggal_kerja()[:4])
    conn = get_db()
    tabel, total_per_bulan, grand_total = _rekap_donasi_donatur(conn, tahun)

    if request.args.get('download'):
        wb = Workbook()
        ws = wb.active
        ws.title = f'Donasi {tahun}'[:31]
        headers = (['Donatur', 'Kategori'] + [BULAN_IND[i][:3] for i in range(1, 13)]
                   + ['Total Setahun', 'Bulan Aktif', 'Rata-rata per Bulan'])
        hdr_font = Font(bold=True, color='FFFFFF', size=10)
        fill = PatternFill('solid', fgColor='1B3B5A')
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
        for r, e in enumerate(tabel, 2):
            ws.cell(row=r, column=1, value=e['nama'])
            ws.cell(row=r, column=2, value=e['sumber_infaq'])
            for i in range(1, 13):
                ws.cell(row=r, column=2 + i, value=e['per_bulan'].get(f"{i:02d}", 0))
            ws.cell(row=r, column=15, value=e['total'])
            ws.cell(row=r, column=16, value=e['bulan_aktif'])
            ws.cell(row=r, column=17, value=round(e['rata2']))
        last = len(tabel) + 2
        bold = Font(bold=True)
        ws.cell(row=last, column=1, value=f'TOTAL {tahun}').font = bold
        for i in range(1, 13):
            ws.cell(row=last, column=2 + i, value=total_per_bulan[f"{i:02d}"]).font = bold
        ws.cell(row=last, column=15, value=grand_total).font = bold
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 24 if col_idx <= 2 else 14
        ws.freeze_panes = 'C2'
        ws.auto_filter.ref = f'A1:{chr(64 + len(headers))}{last}'
        conn.close()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, download_name=f'rekap_donasi_donatur_{tahun}.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_rekap_donasi_donatur.html',
        tabel=tabel, total_per_bulan=total_per_bulan, grand_total=grand_total,
        tahun=tahun, inst=inst, BULAN_IND=BULAN_IND)


@app.route('/admin/laporan/donatur-mustahik')
@admin_required
def laporan_donatur_mustahik():
    """Statistik jumlah & komposisi Donatur dan Mustahik (penerima manfaat) -- bahan
    evaluasi pengurus. Sebelumnya tdk ada laporan khusus utk 2 angka ini."""
    conn = get_db()

    d_total = conn.execute("SELECT COUNT(*) FROM donatur").fetchone()[0]
    d_aktif = conn.execute("SELECT COUNT(*) FROM donatur WHERE aktif=1").fetchone()[0]
    d_rutin = conn.execute("SELECT COUNT(*) FROM donatur WHERE aktif=1 AND aktif_infaq=1").fetchone()[0]

    sumber_rows = conn.execute("""
        SELECT COALESCE(sumber_infaq,'-') as k, COUNT(*) as n FROM donatur
        WHERE aktif=1 GROUP BY k ORDER BY n DESC
    """).fetchall()
    d_sumber = [{'key': r['k'], 'label': LABEL_SUMBER.get(r['k'], r['k']), 'n': r['n'],
                 'pct': round(r['n'] * 100 / d_aktif, 1) if d_aktif else 0} for r in sumber_rows]

    jenis_rows = conn.execute("""
        SELECT COALESCE(jenis,'-') as k, COUNT(*) as n FROM donatur
        WHERE aktif=1 GROUP BY k ORDER BY n DESC
    """).fetchall()
    d_jenis = [{'label': r['k'].replace('_',' ').title(), 'n': r['n'],
                'pct': round(r['n'] * 100 / d_aktif, 1) if d_aktif else 0} for r in jenis_rows]

    d_area_terisi = conn.execute(
        "SELECT COUNT(*) FROM donatur WHERE aktif=1 AND area IS NOT NULL AND area<>''"
    ).fetchone()[0]
    d_area_distinct = conn.execute(
        "SELECT COUNT(DISTINCT area) FROM donatur WHERE aktif=1 AND area IS NOT NULL AND area<>''"
    ).fetchone()[0]
    top_area_rows = conn.execute("""
        SELECT area, COUNT(*) as n FROM donatur
        WHERE aktif=1 AND area IS NOT NULL AND area<>'' GROUP BY area ORDER BY n DESC LIMIT 15
    """).fetchall()
    max_area_n = top_area_rows[0]['n'] if top_area_rows else 1
    d_top_area = [{'label': r['area'], 'n': r['n'],
                   'pct': round(r['n'] * 100 / max_area_n, 1)} for r in top_area_rows]

    m_total = conn.execute("SELECT COUNT(*) FROM penerima_manfaat").fetchone()[0]
    m_aktif = conn.execute("SELECT COUNT(*) FROM penerima_manfaat WHERE aktif=1").fetchone()[0]
    asnaf_rows = conn.execute("""
        SELECT asnaf, COUNT(*) as n FROM penerima_manfaat WHERE aktif=1 GROUP BY asnaf
    """).fetchall()
    asnaf_map = {r['asnaf']: r['n'] for r in asnaf_rows}
    max_asnaf_n = max(asnaf_map.values()) if asnaf_map else 1
    m_asnaf = [{'label': LABEL_ASNAF.get(k, k), 'n': asnaf_map.get(k, 0),
                'pct': round(asnaf_map.get(k, 0) * 100 / max_asnaf_n, 1) if max_asnaf_n else 0}
               for k in LABEL_ASNAF]

    trx_keluar_total = conn.execute("SELECT COUNT(*) FROM transaksi WHERE jenis='keluar'").fetchone()[0]
    trx_keluar_bernama = conn.execute(
        "SELECT COUNT(*) FROM transaksi WHERE jenis='keluar' AND penerima_id IS NOT NULL"
    ).fetchone()[0]

    programs = _program_registry(conn)
    m_per_program = []
    for key, p in programs.items():
        ph = ','.join('?' * len(p['coa_ids']))
        row = conn.execute(
            f"SELECT COUNT(*) as n_trx, COUNT(DISTINCT penerima_id) as n_mustahik, "
            f"COALESCE(SUM(jumlah),0) as total "
            f"FROM transaksi WHERE coa_id IN ({ph}) AND jenis='keluar'",
            p['coa_ids']
        ).fetchone()
        if row['n_trx']:
            m_per_program.append({'label': p['label'], 'code': p['code'], 'jenis_dana': p['jenis_dana'],
                                   'n_trx': row['n_trx'], 'n_mustahik': row['n_mustahik'], 'total': row['total']})
    m_per_program.sort(key=lambda g: -g['total'])

    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_donatur_mustahik.html',
        d_total=d_total, d_aktif=d_aktif, d_nonaktif=d_total - d_aktif,
        d_rutin=d_rutin, d_tidak_rutin=d_aktif - d_rutin,
        d_sumber=d_sumber, d_jenis=d_jenis,
        d_area_terisi=d_area_terisi, d_area_kosong=d_aktif - d_area_terisi,
        d_area_pct=round(d_area_terisi * 100 / d_aktif, 1) if d_aktif else 0,
        d_area_distinct=d_area_distinct, d_top_area=d_top_area,
        m_total=m_total, m_aktif=m_aktif, m_nonaktif=m_total - m_aktif, m_asnaf=m_asnaf,
        m_per_program=m_per_program,
        trx_keluar_total=trx_keluar_total, trx_keluar_bernama=trx_keluar_bernama,
        trx_bernama_pct=round(trx_keluar_bernama * 100 / trx_keluar_total, 1) if trx_keluar_total else 0,
        inst=inst, hari_ini=date.today().strftime('%d %B %Y'))


@app.route('/admin/laporan/kegiatan-penyaluran')
@admin_required
def laporan_kegiatan_penyaluran():
    """Rekap kegiatan penyaluran (nama kegiatan, lokasi, jumlah mustahik, total) --
    hanya transaksi keluar yg field 'Info Kegiatan' -nya sudah diisi saat dicatat."""
    bulan = request.args.get('bulan', get_tanggal_kerja()[:7])
    conn = get_db()
    kegiatan = conn.execute("""
        SELECT t.id, t.tanggal, t.nama_kegiatan, t.lokasi, t.jumlah_mustahik, t.jumlah, t.keterangan,
               c.kode as coa_kode, c.nama as coa_nama, c.jenis_dana
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id=c.id
        WHERE t.jenis='keluar' AND t.nama_kegiatan IS NOT NULL AND strftime('%Y-%m', t.tanggal)=?
        ORDER BY t.tanggal DESC, t.id DESC
    """, (bulan,)).fetchall()

    total_disalurkan = sum(k['jumlah'] for k in kegiatan)
    total_mustahik = sum(k['jumlah_mustahik'] or 0 for k in kegiatan)

    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_kegiatan_penyaluran.html',
        kegiatan=kegiatan, bulan=bulan, total_disalurkan=total_disalurkan,
        total_mustahik=total_mustahik, inst=inst)


# ── Rencana Kerja Tahunan (RKT) ─────────────────────────────────────────────

def _realisasi_per_program(conn, tahun):
    """SUM masuk & keluar per program (key sama dgn _program_registry) sepanjang `tahun` --
    dipakai laporan RKT vs Realisasi. Return {program_key: {'masuk':x,'keluar':y}}."""
    programs = _program_registry(conn)
    hasil = {}
    for key, p in programs.items():
        ph = ','.join('?' * len(p['coa_ids']))
        row = conn.execute(
            f"SELECT COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE 0 END),0) as masuk, "
            f"COALESCE(SUM(CASE WHEN jenis='keluar' THEN jumlah ELSE 0 END),0) as keluar "
            f"FROM transaksi WHERE coa_id IN ({ph}) AND strftime('%Y',tanggal)=?",
            (*p['coa_ids'], tahun)
        ).fetchone()
        hasil[key] = {'masuk': row['masuk'], 'keluar': row['keluar']}
    return hasil, programs


def _renstra_totals(conn, tahun):
    """Total target & realisasi RKT utk `tahun` -- dipakai laporan admin & kartu
    ringkasan RKT di dashboard marketing, spy angkanya selalu konsisten satu sumber.
    Realisasi HANYA dijumlah dr program yg benar2 diberi target (bukan semua program
    yg py aktivitas) -- kalau tidak, persentase bisa meledak krn kebawa realisasi
    program lain yg memang tdk direncanakan sama sekali."""
    realisasi, _ = _realisasi_per_program(conn, tahun)
    target_rows = conn.execute(
        "SELECT program_key, jenis, target_nominal FROM renstra_target WHERE tahun=?", (tahun,)
    ).fetchall()

    target_f = real_f = target_p = real_p = 0
    for r in target_rows:
        real = realisasi.get(r['program_key'], {'masuk': 0, 'keluar': 0})
        if r['jenis'] == 'fundraising':
            target_f += r['target_nominal']
            real_f += real['masuk']
        else:
            target_p += r['target_nominal']
            real_p += real['keluar']

    return {
        'target_f': target_f, 'real_f': real_f,
        'pct_f': round(real_f * 100 / target_f, 1) if target_f else None,
        'target_p': target_p, 'real_p': real_p,
        'pct_p': round(real_p * 100 / target_p, 1) if target_p else None,
    }


@app.route('/admin/renstra', methods=['GET', 'POST'])
@admin_required
def admin_renstra():
    """Susun Rencana Kerja Tahunan: target fundraising & pentasharufan per program,
    plus daftar rencana kegiatan -- acuan evaluasi tahun berikutnya."""
    conn = get_db()
    if request.method == 'POST':
        tahun = request.form.get('tahun', '').strip()
        if not tahun:
            flash('Tahun wajib diisi.', 'danger')
            conn.close()
            return redirect(url_for('admin_renstra'))
        programs = _program_registry(conn)
        for key in programs:
            for jenis in ('fundraising', 'pentasharufan'):
                raw = request.form.get(f'target_{jenis}_{key}', '').strip()
                nominal = parse_jumlah(raw) if raw else None
                if nominal is None:
                    conn.execute("DELETE FROM renstra_target WHERE tahun=? AND program_key=? AND jenis=?",
                                 (tahun, key, jenis))
                else:
                    conn.execute("""
                        INSERT INTO renstra_target (tahun, program_key, jenis, target_nominal)
                        VALUES (?,?,?,?)
                        ON CONFLICT(tahun, program_key, jenis) DO UPDATE SET target_nominal=excluded.target_nominal
                    """, (tahun, key, jenis, nominal))
        conn.commit(); conn.close()
        flash(f'Rencana Kerja Tahunan {tahun} berhasil disimpan.', 'success')
        return redirect(url_for('admin_renstra', tahun=tahun))

    tahun = request.args.get('tahun') or str(int(get_tanggal_kerja()[:4]) + 1)
    programs = _program_registry(conn)
    existing = conn.execute(
        "SELECT program_key, jenis, target_nominal FROM renstra_target WHERE tahun=?", (tahun,)
    ).fetchall()
    target_map = {}
    for r in existing:
        target_map.setdefault(r['program_key'], {})[r['jenis']] = r['target_nominal']

    rows = []
    for key, p in programs.items():
        t = target_map.get(key, {})
        rows.append({
            'key': key, 'label': p['label'], 'code': p['code'], 'jenis_dana': p['jenis_dana'],
            'target_fundraising': t.get('fundraising'),
            'target_pentasharufan': t.get('pentasharufan'),
        })
    rows.sort(key=lambda g: (DANA_TYPES.index(g['jenis_dana']) if g['jenis_dana'] in DANA_TYPES else 99, g['label']))

    kegiatan = conn.execute(
        "SELECT * FROM renstra_kegiatan WHERE tahun=? ORDER BY COALESCE(bulan_rencana,'99'), id", (tahun,)
    ).fetchall()

    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/renstra.html', rows=rows, tahun=tahun, kegiatan=kegiatan,
        programs=programs, dana_types=DANA_TYPES, inst=inst)


@app.route('/admin/renstra/kegiatan/tambah', methods=['POST'])
@admin_required
def renstra_kegiatan_tambah():
    data = request.form
    tahun = data.get('tahun', '').strip()
    if not tahun or not data.get('nama_kegiatan', '').strip():
        flash('Tahun dan Nama Kegiatan wajib diisi.', 'danger')
        return redirect(url_for('admin_renstra', tahun=tahun))
    target_nominal = parse_jumlah(data.get('target_nominal')) or 0
    target_mustahik = None
    if data.get('target_mustahik'):
        try: target_mustahik = int(data['target_mustahik'])
        except ValueError: target_mustahik = None
    conn = get_db()
    conn.execute("""INSERT INTO renstra_kegiatan
        (tahun, program_key, bulan_rencana, nama_kegiatan, lokasi_rencana, target_mustahik, target_nominal, keterangan)
        VALUES (?,?,?,?,?,?,?,?)""",
        (tahun, data.get('program_key') or None, data.get('bulan_rencana') or None,
         data['nama_kegiatan'].strip(), data.get('lokasi_rencana') or None,
         target_mustahik, target_nominal, data.get('keterangan') or None))
    conn.commit(); conn.close()
    flash('Rencana kegiatan ditambahkan.', 'success')
    return redirect(url_for('admin_renstra', tahun=tahun))


@app.route('/admin/renstra/kegiatan/edit/<int:id>', methods=['POST'])
@admin_required
def renstra_kegiatan_edit(id):
    data = request.form
    target_nominal = parse_jumlah(data.get('target_nominal')) or 0
    target_mustahik = None
    if data.get('target_mustahik'):
        try: target_mustahik = int(data['target_mustahik'])
        except ValueError: target_mustahik = None
    conn = get_db()
    conn.execute("""UPDATE renstra_kegiatan SET
        program_key=?, bulan_rencana=?, nama_kegiatan=?, lokasi_rencana=?,
        target_mustahik=?, target_nominal=?, keterangan=?, status=?
        WHERE id=?""",
        (data.get('program_key') or None, data.get('bulan_rencana') or None,
         data.get('nama_kegiatan', '').strip(), data.get('lokasi_rencana') or None,
         target_mustahik, target_nominal, data.get('keterangan') or None,
         data.get('status', 'rencana'), id))
    conn.commit()
    row = conn.execute("SELECT tahun FROM renstra_kegiatan WHERE id=?", (id,)).fetchone()
    conn.close()
    flash('Rencana kegiatan diperbarui.', 'success')
    return redirect(url_for('admin_renstra', tahun=row['tahun'] if row else ''))


@app.route('/admin/renstra/kegiatan/hapus/<int:id>', methods=['POST'])
@admin_required
def renstra_kegiatan_hapus(id):
    conn = get_db()
    row = conn.execute("SELECT tahun FROM renstra_kegiatan WHERE id=?", (id,)).fetchone()
    conn.execute("DELETE FROM renstra_kegiatan WHERE id=?", (id,))
    conn.commit(); conn.close()
    flash('Rencana kegiatan dihapus.', 'success')
    return redirect(url_for('admin_renstra', tahun=row['tahun'] if row else ''))


@app.route('/admin/laporan/renstra')
@admin_required
def laporan_renstra():
    """Evaluasi RKT vs Realisasi -- acuan evaluasi kegiatan tahun berjalan/berikutnya."""
    tahun = request.args.get('tahun', get_tanggal_kerja()[:4])
    conn = get_db()
    realisasi, programs = _realisasi_per_program(conn, tahun)
    target_rows = conn.execute(
        "SELECT program_key, jenis, target_nominal FROM renstra_target WHERE tahun=?", (tahun,)
    ).fetchall()
    target_map = {}
    for r in target_rows:
        target_map.setdefault(r['program_key'], {})[r['jenis']] = r['target_nominal']

    rows = []
    for key, p in programs.items():
        t = target_map.get(key, {})
        tf = t.get('fundraising') or 0
        tp = t.get('pentasharufan') or 0
        rf = realisasi[key]['masuk']
        rp = realisasi[key]['keluar']
        if not (tf or tp or rf or rp):
            continue
        rows.append({
            'label': p['label'], 'code': p['code'], 'jenis_dana': p['jenis_dana'],
            'target_fundraising': tf, 'realisasi_fundraising': rf,
            'pct_fundraising': round(rf * 100 / tf, 1) if tf else None,
            'target_pentasharufan': tp, 'realisasi_pentasharufan': rp,
            'pct_pentasharufan': round(rp * 100 / tp, 1) if tp else None,
        })
    rows.sort(key=lambda g: (DANA_TYPES.index(g['jenis_dana']) if g['jenis_dana'] in DANA_TYPES else 99, g['label']))

    totals = _renstra_totals(conn, tahun)

    kegiatan = conn.execute(
        "SELECT * FROM renstra_kegiatan WHERE tahun=? ORDER BY COALESCE(bulan_rencana,'99'), id", (tahun,)
    ).fetchall()
    n_kegiatan_total = len(kegiatan)
    n_kegiatan_terlaksana = sum(1 for k in kegiatan if k['status'] == 'terlaksana')

    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_renstra.html',
        rows=rows, tahun=tahun, kegiatan=kegiatan,
        total_target_f=totals['target_f'], total_real_f=totals['real_f'],
        total_target_p=totals['target_p'], total_real_p=totals['real_p'],
        pct_f=totals['pct_f'], pct_p=totals['pct_p'],
        n_kegiatan_total=n_kegiatan_total, n_kegiatan_terlaksana=n_kegiatan_terlaksana, inst=inst)


@app.route('/admin/laporan/rkt-cetak')
@admin_required
def laporan_rkt_cetak():
    """Cetak Rencana Kerja Tahunan (RKT) apa adanya -- dokumen rencana murni tanpa
    kolom realisasi, utk pengesahan/arsip di awal tahun sebelum ada data realisasi."""
    tahun = request.args.get('tahun', get_tanggal_kerja()[:4])
    conn = get_db()
    programs = _program_registry(conn)
    target_rows = conn.execute(
        "SELECT program_key, jenis, target_nominal FROM renstra_target WHERE tahun=?", (tahun,)
    ).fetchall()
    target_map = {}
    for r in target_rows:
        target_map.setdefault(r['program_key'], {})[r['jenis']] = r['target_nominal']

    rows = []
    for key, p in programs.items():
        t = target_map.get(key, {})
        tf = t.get('fundraising') or 0
        tp = t.get('pentasharufan') or 0
        if not (tf or tp):
            continue
        rows.append({'label': p['label'], 'code': p['code'], 'jenis_dana': p['jenis_dana'],
                      'target_fundraising': tf, 'target_pentasharufan': tp})
    rows.sort(key=lambda g: (DANA_TYPES.index(g['jenis_dana']) if g['jenis_dana'] in DANA_TYPES else 99, g['label']))

    total_target_f = sum(r['target_fundraising'] for r in rows)
    total_target_p = sum(r['target_pentasharufan'] for r in rows)

    kegiatan = conn.execute(
        "SELECT * FROM renstra_kegiatan WHERE tahun=? ORDER BY COALESCE(bulan_rencana,'99'), id", (tahun,)
    ).fetchall()

    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_rkt_cetak.html',
        rows=rows, tahun=tahun, kegiatan=kegiatan, programs=programs,
        total_target_f=total_target_f, total_target_p=total_target_p, inst=inst)


@app.route('/admin/laporan/laz-pusat')
@admin_required
def laporan_laz_pusat():
    """Isi otomatis template Excel resmi Laporan ZIS ke Yayasan MKU Pusat (6 dari 8 sheet --
    Saldo Rekening & Program Kegiatan tetap manual). Lihat laz_pusat_report.py utk detail."""
    tahun = int(request.args.get('tahun', get_tanggal_kerja()[:4]))
    bulan_sampai = int(request.args.get('bulan_sampai', get_tanggal_kerja()[5:7]))
    conn = get_db()

    if request.args.get('download'):
        buf = laz_pusat_report.isi_template(conn, tahun, bulan_sampai)
        conn.close()
        nama_bulan = BULAN_IND[bulan_sampai]
        return send_file(buf,
            download_name=f'LAPORAN_ZIS_LAZ_MKU_{nama_bulan.upper()}_{tahun}.xlsx',
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    dp = laz_pusat_report.data_penghimpunan(conn, tahun)
    pt = laz_pusat_report.pentasyarufan(conn, tahun)
    total_penghimpunan = sum(dp[m]['zakat_maal'] + dp[m]['zakat_fitrah'] + dp[m]['infaq_bebas']
                              + dp[m]['infaq_terikat'] for m in range(1, bulan_sampai + 1))
    total_penyaluran = sum(sum(pt['bidang'][m].values()) + sum(pt['zakat_asnaf'][m].values())
                            for m in range(1, bulan_sampai + 1))
    kurang_mustahik = laz_pusat_report.hitung_kelengkapan_mustahik(conn, tahun, bulan_sampai)
    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/laporan_laz_pusat.html',
        tahun=tahun, bulan_sampai=bulan_sampai, inst=inst,
        total_penghimpunan=total_penghimpunan, total_penyaluran=total_penyaluran,
        kurang_mustahik=kurang_mustahik, BULAN_IND=BULAN_IND)


@app.route('/admin/laporan/arus-kas')
@admin_required
def laporan_arus_kas():
    bulan    = request.args.get('bulan', get_tanggal_kerja()[:7])
    prev_ld  = _prev_last_day(bulan)
    conn     = get_db()

    saldo_awal_trans = conn.execute("""
        SELECT COALESCE(SUM(CASE WHEN jenis='masuk' THEN jumlah ELSE -jumlah END),0)
        FROM transaksi WHERE tanggal <= ?
    """, (prev_ld,)).fetchone()[0]
    saldo_awal = saldo_awal_trans + sum(get_saldo_awal(conn).values())

    masuk = conn.execute("""
        SELECT c.kode, c.nama, c.jenis_dana, SUM(t.jumlah) as total
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id=c.id
        WHERE t.jenis='masuk' AND strftime('%Y-%m',t.tanggal)=?
        GROUP BY c.id ORDER BY c.kode
    """, (bulan,)).fetchall()

    keluar = conn.execute("""
        SELECT c.kode, c.nama, c.jenis_dana, SUM(t.jumlah) as total
        FROM transaksi t JOIN chart_of_accounts c ON t.coa_id=c.id
        WHERE t.jenis='keluar' AND strftime('%Y-%m',t.tanggal)=?
        GROUP BY c.id ORDER BY c.kode
    """, (bulan,)).fetchall()

    total_masuk  = sum(r['total'] for r in masuk)
    total_keluar = sum(r['total'] for r in keluar)
    conn.close()

    inst = get_instansi()
    return render_template('admin/laporan_arus_kas.html',
        masuk=masuk, keluar=keluar,
        total_masuk=total_masuk, total_keluar=total_keluar,
        saldo_awal=saldo_awal, saldo_akhir=saldo_awal + total_masuk - total_keluar,
        bulan=bulan, inst=inst)

# ── Admin Koleksi ─────────────────────────────────────────────────────────────

@app.route('/admin/koleksi')
@admin_required
def admin_koleksi():
    conn = get_db()
    periode = conn.execute('''
        SELECT bulan,
               COUNT(*) as total,
               SUM(CASE WHEN status='terkumpul' THEN 1 ELSE 0 END) as terkumpul,
               SUM(CASE WHEN status='tidak_ada' THEN 1 ELSE 0 END) as tidak_ada,
               SUM(CASE WHEN status='terjadwal' THEN 1 ELSE 0 END) as terjadwal,
               COALESCE(SUM(CASE WHEN status='terkumpul' THEN jumlah ELSE 0 END),0) as total_nominal
        FROM koleksi_bulanan GROUP BY bulan ORDER BY bulan DESC
    ''').fetchall()
    total_donatur = conn.execute(
        "SELECT COUNT(*) FROM donatur WHERE sumber_infaq IN ('kencleng','kotak_infaq') AND aktif_infaq=1 AND aktif=1"
    ).fetchone()[0]
    conn.close()
    return render_template('admin/koleksi.html', periode=periode,
        total_donatur=total_donatur, bulan_ini=date.today().strftime('%Y-%m'))

@app.route('/admin/koleksi/buka', methods=['POST'])
@admin_required
def admin_koleksi_buka():
    bulan = request.form.get('bulan', date.today().strftime('%Y-%m'))
    created = buka_periode(bulan, session['user_id'])
    if created:
        flash(f'Periode {bulan} dibuka: {created} record koleksi dibuat.', 'success')
    else:
        flash(f'Periode {bulan} sudah ada atau tidak ada donatur aktif.', 'warning')
    return redirect(url_for('admin_koleksi'))

@app.route('/admin/koleksi/<bulan>')
@admin_required
def admin_koleksi_detail(bulan):
    conn = get_db()
    status_filter = request.args.get('status', '')
    area_filter   = request.args.get('area', '')
    query = '''
        SELECT kb.*, d.nama as donatur_nama, d.sumber_infaq, d.area, d.lokasi_nama,
               d.lat, d.lng, u.nama as marketing_nama,
               uk.nama as kunjungi_nama
        FROM koleksi_bulanan kb
        JOIN donatur d ON kb.donatur_id=d.id
        LEFT JOIN users u ON kb.marketing_id=u.id
        LEFT JOIN users uk ON kb.marketing_kunjungi_terakhir=uk.id
        WHERE kb.bulan=?'''
    params = [bulan]
    if status_filter:
        query += ' AND kb.status=?'; params.append(status_filter)
    if area_filter:
        query += ' AND d.area=?'; params.append(area_filter)
    query += ' ORDER BY d.area, d.nama'
    koleksi = conn.execute(query, params).fetchall()
    areas   = conn.execute(
        "SELECT DISTINCT d.area FROM koleksi_bulanan kb JOIN donatur d ON kb.donatur_id=d.id "
        "WHERE kb.bulan=? AND d.area IS NOT NULL ORDER BY d.area", (bulan,)
    ).fetchall()
    stats = conn.execute('''
        SELECT COUNT(*) as total,
               SUM(CASE WHEN status='terkumpul' THEN 1 ELSE 0 END) as terkumpul,
               SUM(CASE WHEN status='tidak_ada' THEN 1 ELSE 0 END) as tidak_ada,
               SUM(CASE WHEN status='terjadwal' THEN 1 ELSE 0 END) as terjadwal,
               COALESCE(SUM(CASE WHEN status='terkumpul' THEN jumlah ELSE 0 END),0) as total_nominal
        FROM koleksi_bulanan WHERE bulan=?
    ''', (bulan,)).fetchone()
    conn.close()
    return render_template('admin/koleksi_detail.html', koleksi=koleksi,
        bulan=bulan, stats=stats, areas=areas,
        status_filter=status_filter, area_filter=area_filter)

@app.route('/admin/koleksi/capaian')
@admin_required
def admin_koleksi_capaian():
    conn = get_db()
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    capaian = conn.execute('''
        SELECT u.nama, u.username,
               COUNT(*) as jumlah_koleksi,
               COALESCE(SUM(kb.jumlah),0) as total_nominal,
               SUM(CASE WHEN d.sumber_infaq='kencleng' THEN 1 ELSE 0 END) as kencleng,
               SUM(CASE WHEN d.sumber_infaq='kotak_infaq' THEN 1 ELSE 0 END) as kotak
        FROM koleksi_bulanan kb
        JOIN users u ON kb.marketing_id=u.id
        JOIN donatur d ON kb.donatur_id=d.id
        WHERE kb.bulan=? AND kb.status='terkumpul'
        GROUP BY u.id ORDER BY total_nominal DESC
    ''', (bulan,)).fetchall()
    conn.close()
    return render_template('admin/koleksi_capaian.html', capaian=capaian, bulan=bulan)

# ── Admin Peta ────────────────────────────────────────────────────────────────

@app.route('/admin/peta')
@admin_required
def admin_peta():
    conn = get_db()
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    donatur = conn.execute('''
        SELECT d.id, d.nama, d.sumber_infaq, d.area, d.desa, d.lokasi_nama, d.lat, d.lng,
               kb.status as koleksi_status, kb.jumlah as koleksi_jumlah
        FROM donatur d
        LEFT JOIN koleksi_bulanan kb ON kb.donatur_id=d.id AND kb.bulan=?
        WHERE d.sumber_infaq IN ('kencleng','kotak_infaq') AND d.aktif=1
              AND d.lat IS NOT NULL AND d.lng IS NOT NULL
        ORDER BY d.area, d.nama
    ''', (bulan,)).fetchall()
    donatur_json = json.dumps([dict(d) for d in donatur])
    conn.close()
    return render_template('admin/peta.html', donatur_json=donatur_json, bulan=bulan)

@app.route('/admin/peta/set-desa', methods=['POST'])
@admin_required
def admin_peta_set_desa():
    data = request.json or {}
    try:
        did = int(data.get('id'))
    except (TypeError, ValueError):
        return jsonify(ok=False, msg='ID tidak valid'), 400
    desa = (data.get('desa') or '').strip()
    conn = get_db()
    conn.execute("UPDATE donatur SET desa=? WHERE id=?", (desa or None, did))
    conn.commit(); conn.close()
    return jsonify(ok=True, id=did, desa=desa)

# ── Master: Users ─────────────────────────────────────────────────────────────

@app.route('/admin/master/users')
@admin_required
def master_users():
    conn = get_db()
    users = conn.execute("SELECT * FROM users ORDER BY role, nama").fetchall()
    conn.close()
    return render_template('admin/master/users.html', users=users)

@app.route('/admin/master/users/tambah', methods=['POST'])
@admin_required
def master_users_tambah():
    data = request.form
    conn = get_db()
    try:
        conn.execute("INSERT INTO users (username,password,nama,role,no_hp) VALUES (?,?,?,?,?)",
                     (data['username'], hash_pw(data['password']), data['nama'],
                      data['role'], data.get('no_hp','')))
        conn.commit(); flash('User berhasil ditambahkan.', 'success')
    except sqlite3.IntegrityError:
        flash('Username sudah dipakai.', 'danger')
    conn.close()
    return redirect(url_for('master_users'))

@app.route('/admin/master/users/edit/<int:id>', methods=['POST'])
@admin_required
def master_users_edit(id):
    data = request.form
    conn = get_db()
    if data.get('password'):
        conn.execute("UPDATE users SET nama=?,role=?,no_hp=?,password=? WHERE id=?",
                     (data['nama'],data['role'],data.get('no_hp',''),hash_pw(data['password']),id))
    else:
        conn.execute("UPDATE users SET nama=?,role=?,no_hp=? WHERE id=?",
                     (data['nama'],data['role'],data.get('no_hp',''),id))
    conn.commit(); conn.close()
    flash('User diperbarui.', 'success')
    return redirect(url_for('master_users'))

@app.route('/admin/master/users/toggle/<int:id>', methods=['POST'])
@admin_required
def master_users_toggle(id):
    conn = get_db()
    conn.execute("UPDATE users SET aktif = CASE WHEN aktif=1 THEN 0 ELSE 1 END WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for('master_users'))

# ── Master: CoA ───────────────────────────────────────────────────────────────

@app.route('/admin/master/coa')
@admin_required
def master_coa():
    conn = get_db()
    coa = conn.execute("SELECT * FROM chart_of_accounts ORDER BY kode").fetchall()
    conn.close()
    return render_template('admin/master/coa.html', coa=coa)

@app.route('/admin/master/coa/tambah', methods=['POST'])
@admin_required
def master_coa_tambah():
    data = request.form
    conn = get_db()
    try:
        conn.execute('''INSERT INTO chart_of_accounts
            (kode,nama,kelompok,jenis_dana,parent_kode,jenis_transaksi) VALUES (?,?,?,?,?,?)''',
            (data['kode'].strip(), data['nama'].strip(), data['kelompok'],
             data.get('jenis_dana') or None, data.get('parent_kode') or None,
             data.get('jenis_transaksi') or None))
        conn.commit(); flash('Akun berhasil ditambahkan.', 'success')
    except sqlite3.IntegrityError:
        flash('Kode akun sudah ada.', 'danger')
    conn.close()
    return redirect(url_for('master_coa'))

@app.route('/admin/master/coa/toggle/<int:id>', methods=['POST'])
@admin_required
def master_coa_toggle(id):
    conn = get_db()
    conn.execute("UPDATE chart_of_accounts SET aktif = CASE WHEN aktif=1 THEN 0 ELSE 1 END WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for('master_coa'))

# ── Master: Donatur ───────────────────────────────────────────────────────────

@app.route('/admin/master/donatur')
@admin_required
def master_donatur():
    conn = get_db()
    q       = request.args.get('q','')
    sumber  = request.args.get('sumber','')
    area    = request.args.get('area','')
    aktif_f = request.args.get('aktif','')
    query   = "SELECT d.*, c.nama AS program_nama FROM donatur d LEFT JOIN chart_of_accounts c ON d.program_id=c.id WHERE 1=1"
    params  = []
    if q:
        query += " AND (d.nama LIKE ? OR d.no_hp LIKE ? OR d.lokasi_nama LIKE ?)";
        params += [f'%{q}%',f'%{q}%',f'%{q}%']
    if sumber:
        query += " AND d.sumber_infaq=?"; params.append(sumber)
    if area:
        query += " AND d.area=?"; params.append(area)
    if aktif_f in ('0','1'):
        query += " AND d.aktif=?"; params.append(int(aktif_f))
    query += " ORDER BY d.sumber_infaq, d.area, d.nama"
    donatur = conn.execute(query, params).fetchall()
    areas   = conn.execute("SELECT nama AS area FROM area WHERE aktif=1 ORDER BY nama").fetchall()
    produk_list = conn.execute(
        "SELECT id, kode, nama, parent_kode FROM chart_of_accounts WHERE parent_kode IN ('4.1','4.2.1','4.4') AND jenis_transaksi='masuk' AND aktif=1 ORDER BY kode"
    ).fetchall()
    conn.close()
    return render_template('admin/master/donatur.html',
        donatur=donatur, areas=areas, produk_list=produk_list, q=q, sumber=sumber, area=area, aktif_f=aktif_f)

@app.route('/admin/master/donatur/tambah', methods=['POST'])
@admin_required
def master_donatur_tambah():
    data = request.form
    lat = lng = None
    gmaps = data.get('gmaps_url','').strip()
    if gmaps:
        lat, lng = parse_gmaps_url(gmaps)
    if lat is None and data.get('lat'):
        try: lat = float(data['lat']); lng = float(data['lng'])
        except (TypeError, ValueError): lat = lng = None
    conn = get_db()
    program_id = int(data['program_id']) if data.get('program_id') else None
    sumber = data.get('sumber_infaq', 'tunai')
    cur = conn.execute("""INSERT INTO donatur
        (nama,nik,no_hp,alamat,jenis,sumber_infaq,area,lokasi_nama,lat,lng,aktif_infaq,program_id)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
        (data['nama'], data.get('nik',''), data.get('no_hp',''), data.get('alamat',''),
         data.get('jenis','perorangan'), sumber,
         data.get('area',''), data.get('lokasi_nama',''),
         lat, lng, 1 if data.get('aktif_infaq') else 0, program_id))
    auto_koleksi_donatur_baru(conn, cur.lastrowid, sumber)
    conn.commit(); conn.close()
    flash('Donatur berhasil ditambahkan.', 'success')
    return redirect(url_for('master_donatur'))

@app.route('/admin/master/donatur/edit/<int:id>', methods=['POST'])
@admin_required
def master_donatur_edit(id):
    data = request.form
    lat = lng = None
    gmaps = data.get('gmaps_url','').strip()
    if gmaps:
        lat, lng = parse_gmaps_url(gmaps)
    if lat is None and data.get('lat'):
        try: lat = float(data['lat']); lng = float(data['lng'])
        except (TypeError, ValueError): lat = lng = None
    conn = get_db()
    existing = conn.execute("SELECT lat,lng FROM donatur WHERE id=?", (id,)).fetchone()
    if lat is None and existing:
        lat, lng = existing['lat'], existing['lng']
    program_id = int(data['program_id']) if data.get('program_id') else None
    conn.execute("""UPDATE donatur SET
        nama=?,nik=?,no_hp=?,alamat=?,jenis=?,sumber_infaq=?,area=?,lokasi_nama=?,lat=?,lng=?,aktif_infaq=?,program_id=?
        WHERE id=?""",
        (data['nama'], data.get('nik',''), data.get('no_hp',''), data.get('alamat',''),
         data.get('jenis','perorangan'), data.get('sumber_infaq','tunai'),
         data.get('area',''), data.get('lokasi_nama',''),
         lat, lng, 1 if data.get('aktif_infaq') else 0, program_id, id))
    conn.commit(); conn.close()
    flash('Donatur diperbarui.', 'success')
    filters = {k[1:]: v for k, v in data.items() if k in ('_q','_sumber','_area','_aktif') and v}
    return redirect(url_for('master_donatur', **filters) + f'#row{id}')

@app.route('/admin/master/donatur/toggle/<int:id>', methods=['POST'])
@admin_required
def master_donatur_toggle(id):
    conn = get_db()
    conn.execute("UPDATE donatur SET aktif = CASE WHEN aktif=1 THEN 0 ELSE 1 END WHERE id=?", (id,))
    conn.commit()
    row = conn.execute("SELECT aktif FROM donatur WHERE id=?", (id,)).fetchone()
    conn.close()
    if row is None:
        return jsonify(ok=False, msg='Donatur tidak ditemukan'), 404
    return jsonify(ok=True, aktif=row['aktif'])

@app.route('/admin/master/donatur/quick/<int:id>', methods=['POST'])
@admin_required
def master_donatur_quick(id):
    ALLOWED = {'nama', 'alamat', 'area', 'desa', 'no_hp', 'sumber_infaq', 'lokasi_nama'}
    data = request.json or {}
    field = data.get('field', '')
    value = data.get('value', '').strip()
    if field not in ALLOWED:
        return jsonify(ok=False, msg='Field tidak valid'), 400
    if field == 'nama' and not value:
        return jsonify(ok=False, msg='Nama wajib diisi'), 400
    conn = get_db()
    conn.execute(f"UPDATE donatur SET {field}=? WHERE id=?", (value or None, id))
    conn.commit(); conn.close()
    return jsonify(ok=True, field=field, value=value)

@app.route('/admin/master/donatur/template')
@admin_required
def donatur_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Donatur'
    headers = ['nama', 'no_hp', 'nik', 'alamat', 'jenis', 'sumber_infaq', 'area', 'lokasi_nama']
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='27AE60')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
    examples = [
        ['Ahmad Fauzi', '6281234567890', '3301010101010001', 'Jl. Mawar No. 5', 'perorangan', 'kencleng', 'Giriwono', 'Rumah Pak Ahmad'],
        ['Toko Berkah', '6289876543210', '', 'Pasar Wonogiri', 'lembaga', 'kotak_infaq', 'Wonokarto', 'Toko Berkah - Pasar'],
    ]
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(italic=True, color='999999')
            cell.border = border
    note = ws.cell(row=4, column=1, value='Catatan: Hapus baris contoh di atas, lalu isi data Anda.')
    note.font = Font(italic=True, color='FF0000')
    note2 = ws.cell(row=5, column=1, value='jenis: perorangan / lembaga | sumber_infaq: tunai / kencleng / kotak_infaq / zakat / infaq_terikat / wakaf')
    note2.font = Font(italic=True, color='666666')
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='template_donatur.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/master/donatur/import', methods=['POST'])
@admin_required
def donatur_import():
    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        flash('Upload file Excel (.xlsx) yang valid.', 'danger')
        return redirect(url_for('master_donatur'))
    try:
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        flash(f'Gagal membaca file: {e}', 'danger')
        return redirect(url_for('master_donatur'))
    if len(rows) < 2:
        flash('File kosong atau hanya berisi header.', 'warning')
        return redirect(url_for('master_donatur'))
    header = [str(h).strip().lower() if h else '' for h in rows[0]]
    required = {'nama'}
    if not required.issubset(set(header)):
        flash('Kolom "nama" wajib ada di header.', 'danger')
        return redirect(url_for('master_donatur'))
    col_map = {h: i for i, h in enumerate(header) if h}
    conn = get_db()
    imported = 0
    skipped = 0
    incomplete = 0
    for row in rows[1:]:
        nama = str(row[col_map['nama']]).strip() if col_map.get('nama') is not None and row[col_map['nama']] else ''
        if not nama or nama.lower() in ('none', 'catatan:', 'catatan'):
            continue
        no_hp = str(row[col_map.get('no_hp', -1)] or '').strip() if 'no_hp' in col_map else ''
        nik = str(row[col_map.get('nik', -1)] or '').strip() if 'nik' in col_map else ''
        alamat = str(row[col_map.get('alamat', -1)] or '').strip() if 'alamat' in col_map else ''
        jenis = str(row[col_map.get('jenis', -1)] or '').strip().lower() if 'jenis' in col_map else 'perorangan'
        if jenis not in ('perorangan', 'lembaga'):
            jenis = 'perorangan'
        sumber = str(row[col_map.get('sumber_infaq', -1)] or '').strip().lower() if 'sumber_infaq' in col_map else 'tunai'
        if sumber not in ('tunai', 'kencleng', 'kotak_infaq', 'zakat', 'infaq_terikat', 'wakaf'):
            sumber = 'tunai'
        area_val = str(row[col_map.get('area', -1)] or '').strip() if 'area' in col_map else ''
        lokasi = str(row[col_map.get('lokasi_nama', -1)] or '').strip() if 'lokasi_nama' in col_map else ''
        existing = conn.execute("SELECT id FROM donatur WHERE LOWER(TRIM(nama))=LOWER(?)", (nama,)).fetchone()
        if existing:
            skipped += 1
            continue
        if not no_hp or not alamat or not area_val:
            incomplete += 1
        conn.execute("""INSERT INTO donatur (nama,no_hp,nik,alamat,jenis,sumber_infaq,area,lokasi_nama)
                        VALUES (?,?,?,?,?,?,?,?)""",
                     (nama, no_hp, nik, alamat, jenis, sumber, area_val, lokasi))
        imported += 1
    conn.commit()
    conn.close()
    msg = f'Import selesai: {imported} donatur ditambahkan, {skipped} duplikat dilewati.'
    if incomplete:
        msg += f' ({incomplete} dari {imported} data tidak lengkap — no_hp/alamat/area kosong)'
    flash(msg, 'success' if not incomplete else 'warning')
    return redirect(url_for('master_donatur'))

# ── Kelompok Penyaluran ───────────────────────────────────────────────────────

@app.route('/admin/kelompok')
@admin_required
def kelompok_list():
    conn = get_db()
    kelompok = conn.execute("""
        SELECT k.*, c.kode AS coa_kode, c.nama AS coa_nama,
               (SELECT COUNT(*) FROM kelompok_penyaluran_anggota a
                WHERE a.kelompok_id=k.id AND a.aktif=1) AS jml_anggota
        FROM kelompok_penyaluran k
        LEFT JOIN chart_of_accounts c ON k.coa_id=c.id
        ORDER BY k.nama
    """).fetchall()
    coa_list = conn.execute(
        "SELECT id, kode, nama FROM chart_of_accounts WHERE jenis_transaksi='keluar' AND aktif=1 ORDER BY kode"
    ).fetchall()
    conn.close()
    return render_template('admin/kelompok.html', kelompok=kelompok, coa_list=coa_list)

@app.route('/admin/kelompok/tambah', methods=['POST'])
@admin_required
def kelompok_tambah():
    data = request.form
    conn = get_db()
    conn.execute("""INSERT INTO kelompok_penyaluran
        (nama, coa_id, pakai_token, template_pesan, template_pesan_prepaid)
        VALUES (?,?,?,?,?)""",
        (data['nama'].strip(), int(data['coa_id']), 1 if data.get('pakai_token') else 0,
         data.get('template_pesan', '').strip(), data.get('template_pesan_prepaid', '').strip()))
    conn.commit(); conn.close()
    flash('Kelompok penyaluran berhasil dibuat.', 'success')
    return redirect(url_for('kelompok_list'))

@app.route('/admin/kelompok/<int:id>')
@admin_required
def kelompok_detail(id):
    conn = get_db()
    kelompok = conn.execute("""
        SELECT k.*, c.kode AS coa_kode, c.nama AS coa_nama
        FROM kelompok_penyaluran k LEFT JOIN chart_of_accounts c ON k.coa_id=c.id
        WHERE k.id=?
    """, (id,)).fetchone()
    if not kelompok:
        conn.close()
        flash('Kelompok tidak ditemukan.', 'danger')
        return redirect(url_for('kelompok_list'))
    anggota = conn.execute("""
        SELECT a.*, p.nama AS penerima_nama, p.no_hp, p.alamat
        FROM kelompok_penyaluran_anggota a JOIN penerima_manfaat p ON a.penerima_id=p.id
        WHERE a.kelompok_id=? ORDER BY a.urutan, p.nama
    """, (id,)).fetchall()
    penerima_tersedia = conn.execute("""
        SELECT id, nama FROM penerima_manfaat
        WHERE aktif=1 AND id NOT IN
            (SELECT penerima_id FROM kelompok_penyaluran_anggota WHERE kelompok_id=?)
        ORDER BY nama
    """, (id,)).fetchall()
    periode = conn.execute("""
        SELECT b.bulan,
               COUNT(*) AS total,
               SUM(CASE WHEN b.status='tersalur' THEN 1 ELSE 0 END) AS tersalur,
               SUM(CASE WHEN b.status='tersalur' THEN b.jumlah ELSE 0 END) AS total_nominal
        FROM kelompok_penyaluran_bulanan b
        JOIN kelompok_penyaluran_anggota a ON b.anggota_id=a.id
        WHERE a.kelompok_id=?
        GROUP BY b.bulan ORDER BY b.bulan DESC
    """, (id,)).fetchall()
    conn.close()
    return render_template('admin/kelompok_detail.html', kelompok=kelompok, anggota=anggota,
                            penerima_tersedia=penerima_tersedia, periode=periode,
                            bulan_ini=date.today().strftime('%Y-%m'))

@app.route('/admin/kelompok/<int:id>/anggota/tambah', methods=['POST'])
@admin_required
def kelompok_anggota_tambah(id):
    data = request.form
    conn = get_db()
    try:
        conn.execute("""INSERT INTO kelompok_penyaluran_anggota (kelompok_id, penerima_id, tipe)
            VALUES (?,?,?)""", (id, int(data['penerima_id']), data.get('tipe') or None))
        conn.commit()
        flash('Anggota ditambahkan ke kelompok.', 'success')
    except sqlite3.IntegrityError:
        flash('Penerima ini sudah jadi anggota kelompok.', 'warning')
    conn.close()
    return redirect(url_for('kelompok_detail', id=id))

@app.route('/admin/kelompok/<int:id>/anggota/<int:anggota_id>/toggle', methods=['POST'])
@admin_required
def kelompok_anggota_toggle(id, anggota_id):
    conn = get_db()
    conn.execute("""UPDATE kelompok_penyaluran_anggota SET aktif = CASE WHEN aktif=1 THEN 0 ELSE 1 END
        WHERE id=? AND kelompok_id=?""", (anggota_id, id))
    conn.commit(); conn.close()
    return redirect(url_for('kelompok_detail', id=id))

@app.route('/admin/kelompok/<int:id>/buka', methods=['POST'])
@admin_required
def kelompok_buka_periode(id):
    bulan = request.form.get('bulan', '').strip()
    if not bulan:
        flash('Bulan wajib diisi.', 'danger')
        return redirect(url_for('kelompok_detail', id=id))
    if not re.fullmatch(r'\d{4}-\d{2}', bulan):
        flash('Format bulan tidak valid.', 'danger')
        return redirect(url_for('kelompok_detail', id=id))
    conn = get_db()
    anggota = conn.execute(
        "SELECT id FROM kelompok_penyaluran_anggota WHERE kelompok_id=? AND aktif=1", (id,)
    ).fetchall()
    created = 0
    for a in anggota:
        prefill = conn.execute("""
            SELECT jumlah FROM kelompok_penyaluran_bulanan
            WHERE anggota_id=? AND bulan < ? ORDER BY bulan DESC LIMIT 1
        """, (a['id'], bulan)).fetchone()
        try:
            conn.execute("""INSERT INTO kelompok_penyaluran_bulanan (anggota_id, bulan, jumlah)
                VALUES (?,?,?)""", (a['id'], bulan, prefill['jumlah'] if prefill else None))
            created += 1
        except sqlite3.IntegrityError:
            pass
    conn.commit(); conn.close()
    flash(f'Periode {bulan} dibuka, {created} baris dibuat.', 'success')
    return redirect(url_for('kelompok_bulanan_detail', id=id, bulan=bulan))

@app.route('/admin/kelompok/<int:id>/<bulan>')
@admin_required
def kelompok_bulanan_detail(id, bulan):
    conn = get_db()
    kelompok = conn.execute("SELECT * FROM kelompok_penyaluran WHERE id=?", (id,)).fetchone()
    if not kelompok:
        conn.close()
        flash('Kelompok tidak ditemukan.', 'danger')
        return redirect(url_for('kelompok_list'))
    baris = conn.execute("""
        SELECT b.*, p.nama AS penerima_nama, p.no_hp, a.tipe
        FROM kelompok_penyaluran_bulanan b
        JOIN kelompok_penyaluran_anggota a ON b.anggota_id=a.id
        JOIN penerima_manfaat p ON a.penerima_id=p.id
        WHERE a.kelompok_id=? AND b.bulan=?
        ORDER BY a.urutan, p.nama
    """, (id, bulan)).fetchall()
    conn.close()
    return render_template('admin/kelompok_bulanan.html', kelompok=kelompok, bulan=bulan, baris=baris)

def kirim_notifikasi_penyaluran(conn, bulanan_id):
    """Kirim WA notifikasi utk 1 baris kelompok_penyaluran_bulanan, lalu update
    status kirimnya di DB (tidak commit — caller yang commit)."""
    row = conn.execute("""
        SELECT b.id, b.jumlah, b.token, b.bulan, p.nama AS penerima_nama, p.no_hp, a.tipe,
               k.pakai_token, k.template_pesan, k.template_pesan_prepaid
        FROM kelompok_penyaluran_bulanan b
        JOIN kelompok_penyaluran_anggota a ON b.anggota_id=a.id
        JOIN penerima_manfaat p ON a.penerima_id=p.id
        JOIN kelompok_penyaluran k ON a.kelompok_id=k.id
        WHERE b.id=?
    """, (bulanan_id,)).fetchone()
    if not row:
        return
    if not row['no_hp']:
        conn.execute("UPDATE kelompok_penyaluran_bulanan SET wa_status='gagal', wa_error=? WHERE id=?",
                     ('Nomor HP kosong', bulanan_id))
        return
    template = row['template_pesan']
    if row['pakai_token'] and row['tipe'] == 'prepaid':
        template = row['template_pesan_prepaid']
    pesan = render_pesan(template or '', nama=row['penerima_nama'], bulan=format_bulan(row['bulan']),
                          nominal=format_rupiah(row['jumlah']), token=row['token'] or '')
    ok, error = kirim_wa(row['no_hp'], pesan)
    if ok:
        conn.execute("""UPDATE kelompok_penyaluran_bulanan
            SET wa_status='terkirim', wa_error=NULL, wa_sent_at=datetime('now','localtime') WHERE id=?""",
            (bulanan_id,))
    else:
        conn.execute("UPDATE kelompok_penyaluran_bulanan SET wa_status='gagal', wa_error=? WHERE id=?",
                     (error, bulanan_id))

@app.route('/admin/kelompok/<int:id>/<bulan>/simpan', methods=['POST'])
@admin_required
def kelompok_bulanan_simpan(id, bulan):
    conn = get_db()
    kelompok = conn.execute("SELECT * FROM kelompok_penyaluran WHERE id=?", (id,)).fetchone()
    if not kelompok:
        conn.close()
        flash('Kelompok tidak ditemukan.', 'danger')
        return redirect(url_for('kelompok_list'))
    baris = conn.execute("""
        SELECT b.id, b.status, b.transaksi_id, b.wa_status, a.penerima_id
        FROM kelompok_penyaluran_bulanan b
        JOIN kelompok_penyaluran_anggota a ON b.anggota_id=a.id
        WHERE a.kelompok_id=? AND b.bulan=?
    """, (id, bulan)).fetchall()
    tanggal = get_tanggal_kerja()
    disimpan = 0
    perlu_kirim = []
    for b in baris:
        jumlah = parse_jumlah(request.form.get(f'jumlah_{b["id"]}'))
        token = request.form.get(f'token_{b["id"]}', '').strip()
        if jumlah is None:
            continue
        if b['status'] == 'tersalur' and b['transaksi_id']:
            # Sudah pernah disimpan (mis. koreksi nominal) — update transaksi
            # yang ada, jangan bikin baris baru supaya buku besar tidak dobel.
            cur = conn.execute("UPDATE transaksi SET jumlah=? WHERE id=?", (jumlah, b['transaksi_id']))
            if cur.rowcount:
                trx_id = b['transaksi_id']
            else:
                # Transaksi lama sudah dihapus manual (mis. dari halaman
                # Transaksi) — jangan diam-diam gagal, bikin transaksi baru.
                trx_id, _ = insert_transaksi(
                    conn, tanggal, 'keluar', kelompok['coa_id'], None, b['penerima_id'], jumlah,
                    f"Penyaluran {kelompok['nama']} – {format_bulan(bulan)}", session['user_id'],
                )
        else:
            trx_id, _ = insert_transaksi(
                conn, tanggal, 'keluar', kelompok['coa_id'], None, b['penerima_id'], jumlah,
                f"Penyaluran {kelompok['nama']} – {format_bulan(bulan)}", session['user_id'],
            )
        conn.execute("""UPDATE kelompok_penyaluran_bulanan
            SET jumlah=?, token=?, status='tersalur', transaksi_id=? WHERE id=?""",
            (jumlah, token or None, trx_id, b['id']))
        disimpan += 1
        if b['wa_status'] != 'terkirim':
            perlu_kirim.append(b['id'])
    conn.commit()
    # ponytail: kirim WA setelah transaksi ter-commit, satu commit per baris --
    # supaya gateway yg macet tidak menahan lock DB / menggagalkan seluruh batch
    # keuangan yg sudah tersimpan. Sinkron dalam request (blocking ~1-2 detik x
    # jumlah anggota) aman utk skala saat ini (~40 anggota/bulan, sekali sebulan).
    # Kalau kelompok makin besar, pindah ke background job sebelum ini menabrak
    # timeout gunicorn (timeout dinaikkan ke 180s sbg jaring pengaman, bukan solusi permanen).
    for bid in perlu_kirim:
        kirim_notifikasi_penyaluran(conn, bid)
        conn.commit()
        time.sleep(2)
    conn.close()
    flash(f'{disimpan} penyaluran disimpan.', 'success')
    return redirect(url_for('kelompok_bulanan_detail', id=id, bulan=bulan))

@app.route('/admin/kelompok/bulanan/<int:bulanan_id>/kirim-ulang', methods=['POST'])
@admin_required
def kelompok_bulanan_kirim_ulang(bulanan_id):
    conn = get_db()
    row = conn.execute("""
        SELECT b.bulan, a.kelompok_id FROM kelompok_penyaluran_bulanan b
        JOIN kelompok_penyaluran_anggota a ON b.anggota_id=a.id
        WHERE b.id=?
    """, (bulanan_id,)).fetchone()
    if not row:
        conn.close()
        flash('Data tidak ditemukan.', 'danger')
        return redirect(url_for('kelompok_list'))
    kirim_notifikasi_penyaluran(conn, bulanan_id)
    conn.commit(); conn.close()
    return redirect(url_for('kelompok_bulanan_detail', id=row['kelompok_id'], bulan=row['bulan']))

# ── Master: Penerima Manfaat ──────────────────────────────────────────────────

@app.route('/admin/master/penerima')
@admin_required
def master_penerima():
    conn = get_db()
    q = request.args.get('q',''); asnaf = request.args.get('asnaf','')
    query = "SELECT * FROM penerima_manfaat WHERE 1=1"; params = []
    if q:
        query += " AND (nama LIKE ? OR nik LIKE ? OR no_hp LIKE ?)"; params += [f'%{q}%']*3
    if asnaf:
        query += " AND asnaf=?"; params.append(asnaf)
    penerima = conn.execute(query + " ORDER BY nama", params).fetchall()
    conn.close()
    return render_template('admin/master/penerima.html', penerima=penerima, q=q, asnaf=asnaf)

@app.route('/admin/master/penerima/tambah', methods=['POST'])
@admin_required
def master_penerima_tambah():
    data = request.form
    conn = get_db()
    conn.execute("INSERT INTO penerima_manfaat (nama,nik,no_hp,alamat,asnaf,keterangan) VALUES (?,?,?,?,?,?)",
                 (data['nama'],data.get('nik',''),data.get('no_hp',''),
                  data.get('alamat',''),data.get('asnaf',''),data.get('keterangan','')))
    conn.commit(); conn.close()
    flash('Penerima manfaat berhasil ditambahkan.', 'success')
    return redirect(url_for('master_penerima'))

@app.route('/admin/master/penerima/edit/<int:id>', methods=['POST'])
@admin_required
def master_penerima_edit(id):
    data = request.form
    conn = get_db()
    conn.execute("UPDATE penerima_manfaat SET nama=?,nik=?,no_hp=?,alamat=?,asnaf=?,keterangan=? WHERE id=?",
                 (data['nama'],data.get('nik',''),data.get('no_hp',''),
                  data.get('alamat',''),data.get('asnaf',''),data.get('keterangan',''),id))
    conn.commit(); conn.close()
    flash('Penerima manfaat diperbarui.', 'success')
    return redirect(url_for('master_penerima'))

@app.route('/admin/master/penerima/toggle/<int:id>', methods=['POST'])
@admin_required
def master_penerima_toggle(id):
    conn = get_db()
    conn.execute("UPDATE penerima_manfaat SET aktif = CASE WHEN aktif=1 THEN 0 ELSE 1 END WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for('master_penerima'))

@app.route('/admin/master/penerima/template')
@admin_required
def penerima_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Penerima Manfaat'
    headers = ['nama', 'nik', 'no_hp', 'alamat', 'asnaf', 'keterangan']
    hdr_font = Font(bold=True, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='E74C3C')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
    examples = [
        ['Siti Aminah', '3301010203040005', '6281234000111', 'Dusun Bakaran RT 02/05', 'fakir', 'Janda, 3 anak'],
        ['Ahmad Soleh', '3301010203040006', '6289876000222', 'Ds. Krisak RT 01/03', 'miskin', 'Buruh tani'],
    ]
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(italic=True, color='999999')
            cell.border = border
    note = ws.cell(row=4, column=1, value='Catatan: Hapus baris contoh di atas, lalu isi data Anda.')
    note.font = Font(italic=True, color='FF0000')
    note2 = ws.cell(row=5, column=1, value='asnaf: fakir / miskin / amil / muallaf / riqab / gharim / fisabilillah / ibnu_sabil')
    note2.font = Font(italic=True, color='666666')
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='template_penerima_manfaat.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/master/penerima/import', methods=['POST'])
@admin_required
def penerima_import():
    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        flash('Upload file Excel (.xlsx) yang valid.', 'danger')
        return redirect(url_for('master_penerima'))
    try:
        wb = load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        flash(f'Gagal membaca file: {e}', 'danger')
        return redirect(url_for('master_penerima'))
    if len(rows) < 2:
        flash('File kosong atau hanya berisi header.', 'warning')
        return redirect(url_for('master_penerima'))
    header = [str(h).strip().lower() if h else '' for h in rows[0]]
    if 'nama' not in header:
        flash('Kolom "nama" wajib ada di header.', 'danger')
        return redirect(url_for('master_penerima'))
    col_map = {h: i for i, h in enumerate(header) if h}
    valid_asnaf = {'fakir', 'miskin', 'amil', 'muallaf', 'riqab', 'gharim', 'fisabilillah', 'ibnu_sabil'}
    conn = get_db()
    imported = 0
    skipped = 0
    incomplete = 0
    for row in rows[1:]:
        nama = str(row[col_map['nama']]).strip() if col_map.get('nama') is not None and row[col_map['nama']] else ''
        if not nama or nama.lower() in ('none', 'catatan:', 'catatan'):
            continue
        nik = str(row[col_map.get('nik', -1)] or '').strip() if 'nik' in col_map else ''
        no_hp = str(row[col_map.get('no_hp', -1)] or '').strip() if 'no_hp' in col_map else ''
        alamat = str(row[col_map.get('alamat', -1)] or '').strip() if 'alamat' in col_map else ''
        asnaf = str(row[col_map.get('asnaf', -1)] or '').strip().lower() if 'asnaf' in col_map else ''
        if asnaf not in valid_asnaf:
            asnaf = ''
        keterangan = str(row[col_map.get('keterangan', -1)] or '').strip() if 'keterangan' in col_map else ''
        existing = conn.execute("SELECT id FROM penerima_manfaat WHERE LOWER(TRIM(nama))=LOWER(?)", (nama,)).fetchone()
        if existing:
            skipped += 1
            continue
        if not alamat or not asnaf or not no_hp:
            incomplete += 1
        conn.execute("INSERT INTO penerima_manfaat (nama,nik,no_hp,alamat,asnaf,keterangan) VALUES (?,?,?,?,?,?)",
                     (nama, nik, no_hp, alamat, asnaf, keterangan))
        imported += 1
    conn.commit()
    conn.close()
    msg = f'Import selesai: {imported} penerima ditambahkan, {skipped} duplikat dilewati.'
    if incomplete:
        msg += f' ({incomplete} dari {imported} data tidak lengkap — alamat/asnaf/no_hp kosong)'
    flash(msg, 'success' if not incomplete else 'warning')
    return redirect(url_for('master_penerima'))

# ── Master: Area ─────────────────────────────────────────────────────────────

@app.route('/admin/master/area')
@admin_required
def master_area():
    conn = get_db()
    areas = conn.execute("SELECT a.*, (SELECT COUNT(*) FROM donatur d WHERE d.area=a.nama) AS jml_donatur FROM area a ORDER BY a.nama").fetchall()
    conn.close()
    return render_template('admin/master/area.html', areas=areas)

@app.route('/admin/master/area/tambah', methods=['POST'])
@admin_required
def master_area_tambah():
    nama = request.form.get('nama', '').strip()
    if not nama:
        flash('Nama area wajib diisi.', 'danger')
        return redirect(url_for('master_area'))
    conn = get_db()
    existing = conn.execute("SELECT id FROM area WHERE nama=?", (nama,)).fetchone()
    if existing:
        flash('Area sudah ada.', 'warning')
    else:
        conn.execute("INSERT INTO area (nama) VALUES (?)", (nama,))
        conn.commit()
        flash(f'Area "{nama}" ditambahkan.', 'success')
    conn.close()
    return redirect(url_for('master_area'))

@app.route('/admin/master/area/edit/<int:id>', methods=['POST'])
@admin_required
def master_area_edit(id):
    nama = request.form.get('nama', '').strip()
    if not nama:
        flash('Nama area wajib diisi.', 'danger')
        return redirect(url_for('master_area'))
    conn = get_db()
    old = conn.execute("SELECT nama FROM area WHERE id=?", (id,)).fetchone()
    dup = conn.execute("SELECT id FROM area WHERE nama=? AND id!=?", (nama, id)).fetchone()
    if dup:
        flash('Nama area sudah dipakai.', 'warning')
    else:
        conn.execute("UPDATE area SET nama=? WHERE id=?", (nama, id))
        if old and old['nama']:
            conn.execute("UPDATE donatur SET area=? WHERE area=?", (nama, old['nama']))
        conn.commit()
        flash(f'Area diperbarui menjadi "{nama}".', 'success')
    conn.close()
    return redirect(url_for('master_area'))

@app.route('/admin/master/area/toggle/<int:id>', methods=['POST'])
@admin_required
def master_area_toggle(id):
    conn = get_db()
    conn.execute("UPDATE area SET aktif = CASE WHEN aktif=1 THEN 0 ELSE 1 END WHERE id=?", (id,))
    conn.commit(); conn.close()
    return redirect(url_for('master_area'))

@app.route('/admin/master/area/hapus/<int:id>', methods=['POST'])
@admin_required
def master_area_hapus(id):
    conn = get_db()
    cnt = conn.execute("SELECT COUNT(*) FROM donatur WHERE area=(SELECT nama FROM area WHERE id=?)", (id,)).fetchone()[0]
    if cnt > 0:
        flash(f'Tidak bisa hapus — masih ada {cnt} donatur di area ini.', 'danger')
    else:
        conn.execute("DELETE FROM area WHERE id=?", (id,))
        conn.commit()
        flash('Area dihapus.', 'success')
    conn.close()
    return redirect(url_for('master_area'))

# ── Master: Instansi ─────────────────────────────────────────────────────────

@app.route('/admin/master/instansi')
@admin_required
def master_instansi():
    conn = get_db()
    inst = get_instansi(conn)
    conn.close()
    return render_template('admin/master/instansi.html', inst=inst)

@app.route('/admin/master/instansi/simpan', methods=['POST'])
@admin_required
def master_instansi_simpan():
    data = request.form
    conn = get_db()
    conn.execute("""UPDATE instansi SET
        nama=?, nama_lembaga=?, alamat=?, telepon=?, email=?, website=?,
        ketua=?, bendahara=?, sekretaris=?, no_izin=?,
        updated_at=datetime('now','localtime')
        WHERE id=1""",
        (data.get('nama','').strip(), data.get('nama_lembaga','').strip(),
         data.get('alamat','').strip(), data.get('telepon','').strip(),
         data.get('email','').strip(), data.get('website','').strip(),
         data.get('ketua','').strip(), data.get('bendahara','').strip(),
         data.get('sekretaris','').strip(), data.get('no_izin','').strip()))
    conn.commit(); conn.close()
    flash('Data instansi berhasil diperbarui.', 'success')
    return redirect(url_for('master_instansi'))

# ── Master: Saldo Awal ───────────────────────────────────────────────────────

@app.route('/admin/master/saldo-awal')
@admin_required
def master_saldo_awal():
    conn = get_db()
    saldo = get_saldo_awal(conn)
    ket_rows = conn.execute("SELECT jenis_dana, keterangan FROM saldo_awal").fetchall()
    keterangan = {r['jenis_dana']: r['keterangan'] for r in ket_rows}
    conn.close()
    return render_template('admin/master/saldo_awal.html',
        saldo=saldo, keterangan=keterangan, dana_types=DANA_TYPES)

@app.route('/admin/master/saldo-awal/simpan', methods=['POST'])
@admin_required
def master_saldo_awal_simpan():
    data = request.form
    conn = get_db()
    for dana in DANA_TYPES:
        jumlah = parse_jumlah(data.get(f'jumlah_{dana}')) or 0
        ket = data.get(f'keterangan_{dana}', '').strip()
        conn.execute("""
            INSERT INTO saldo_awal (jenis_dana, jumlah, keterangan, updated_at)
            VALUES (?,?,?,datetime('now','localtime'))
            ON CONFLICT(jenis_dana) DO UPDATE SET
                jumlah=excluded.jumlah, keterangan=excluded.keterangan,
                updated_at=excluded.updated_at
        """, (dana, jumlah, ket))
    conn.commit(); conn.close()
    flash('Saldo awal berhasil disimpan.', 'success')
    return redirect(url_for('master_saldo_awal'))

# ── Admin Jurnal (Non-Tunai) ─────────────────────────────────────────────────

@app.route('/admin/jurnal')
@admin_required
def admin_jurnal():
    conn = get_db()
    bulan = request.args.get('bulan', get_tanggal_kerja()[:7])
    jurnal = conn.execute('''
        SELECT j.*, cd.kode as debit_kode, cd.nama as debit_nama,
               ck.kode as kredit_kode, ck.nama as kredit_nama, u.nama as petugas
        FROM jurnal j
        LEFT JOIN chart_of_accounts cd ON j.debit_coa_id=cd.id
        LEFT JOIN chart_of_accounts ck ON j.kredit_coa_id=ck.id
        LEFT JOIN users u ON j.user_id=u.id
        WHERE strftime('%Y-%m',j.tanggal)=?
        ORDER BY j.tanggal DESC, j.created_at DESC
    ''', (bulan,)).fetchall()
    coa_all = conn.execute(
        "SELECT id, kode, nama, kelompok, jenis_dana FROM chart_of_accounts WHERE aktif=1 ORDER BY kode"
    ).fetchall()
    conn.close()
    return render_template('admin/jurnal.html', jurnal=jurnal, coa_all=coa_all, bulan=bulan)

@app.route('/admin/jurnal/tambah', methods=['POST'])
@admin_required
def admin_jurnal_tambah():
    data = request.form
    try:
        debit_coa_id = int(data['debit_coa_id'])
        kredit_coa_id = int(data['kredit_coa_id'])
    except (KeyError, ValueError):
        flash('Akun debit/kredit belum dipilih.', 'danger')
        return redirect(url_for('admin_jurnal'))
    if debit_coa_id == kredit_coa_id:
        flash('Akun debit dan kredit tidak boleh sama.', 'danger')
        return redirect(url_for('admin_jurnal'))
    jumlah = parse_jumlah(data.get('jumlah'))
    if jumlah is None:
        flash('Jumlah tidak valid — harus angka lebih dari 0.', 'danger')
        return redirect(url_for('admin_jurnal'))
    tanggal = data['tanggal']
    keterangan = data.get('keterangan', '')
    no_bukti = data.get('no_bukti', '').strip()
    client_uuid = data.get('client_uuid') or None

    conn = get_db()
    if client_uuid:
        existing = conn.execute("SELECT id FROM jurnal WHERE client_uuid=?", (client_uuid,)).fetchone()
        if existing:
            conn.close()
            flash('Jurnal ini sudah tercatat sebelumnya — tidak dicatat ganda.', 'warning')
            return redirect(url_for('admin_jurnal'))

    cur = conn.execute("""INSERT INTO jurnal
        (tanggal, no_bukti, keterangan, debit_coa_id, kredit_coa_id, jumlah, user_id, client_uuid)
        VALUES (?,?,?,?,?,?,?,?)""",
        (tanggal, no_bukti, keterangan, debit_coa_id, kredit_coa_id, jumlah,
         session['user_id'], client_uuid))
    jurnal_id = cur.lastrowid

    debit_coa = conn.execute("SELECT jenis_dana, jenis_transaksi FROM chart_of_accounts WHERE id=?",
                              (debit_coa_id,)).fetchone()
    kredit_coa = conn.execute("SELECT jenis_dana, jenis_transaksi FROM chart_of_accounts WHERE id=?",
                               (kredit_coa_id,)).fetchone()

    conn.execute("""INSERT INTO transaksi
        (tanggal, jenis, jenis_dana, coa_id, jumlah, keterangan, user_id, jurnal_id)
        VALUES (?,?,?,?,?,?,?,?)""",
        (tanggal, 'masuk', debit_coa['jenis_dana'] if debit_coa else None,
         debit_coa_id, jumlah, f'[Jurnal] {keterangan}', session['user_id'], jurnal_id))

    conn.execute("""INSERT INTO transaksi
        (tanggal, jenis, jenis_dana, coa_id, jumlah, keterangan, user_id, jurnal_id)
        VALUES (?,?,?,?,?,?,?,?)""",
        (tanggal, 'keluar', kredit_coa['jenis_dana'] if kredit_coa else None,
         kredit_coa_id, jumlah, f'[Jurnal] {keterangan}', session['user_id'], jurnal_id))

    conn.commit(); conn.close()
    flash('Jurnal berhasil dicatat.', 'success')
    return redirect(url_for('admin_jurnal'))

@app.route('/admin/jurnal/hapus/<int:id>', methods=['POST'])
@admin_required
def admin_jurnal_hapus(id):
    conn = get_db()
    conn.execute("DELETE FROM transaksi WHERE jurnal_id=?", (id,))
    conn.execute("DELETE FROM jurnal WHERE id=?", (id,))
    conn.commit(); conn.close()
    flash('Jurnal dan transaksi terkait dihapus.', 'warning')
    return redirect(url_for('admin_jurnal'))

# ── Backup Database ──────────────────────────────────────────────────────────

BACKUP_DIR = os.path.join('data', 'backups')

def get_backup_list():
    if not os.path.exists(BACKUP_DIR):
        return []
    files = glob_mod.glob(os.path.join(BACKUP_DIR, 'backup_*.db'))
    backups = []
    for f in sorted(files, reverse=True):
        fname = os.path.basename(f)
        size = os.path.getsize(f)
        mtime = datetime.fromtimestamp(os.path.getmtime(f))
        label = 'Manual'
        if '_auto_' in fname:
            label = 'Otomatis'
        backups.append({'filename': fname, 'size': size, 'date': mtime, 'label': label})
    return backups

def create_backup(prefix='manual'):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f'backup_{prefix}_{ts}.db'
    dst = os.path.join(BACKUP_DIR, fname)
    src = DB_PATH
    conn = sqlite3.connect(src)
    bak = sqlite3.connect(dst)
    conn.backup(bak)
    bak.close()
    conn.close()
    return fname

@app.route('/admin/backup')
@admin_required
def admin_backup():
    backups = get_backup_list()
    total_size = sum(b['size'] for b in backups)
    conn = get_db()
    stats = {
        'donatur': conn.execute("SELECT COUNT(*) FROM donatur").fetchone()[0],
        'penerima': conn.execute("SELECT COUNT(*) FROM penerima_manfaat").fetchone()[0],
        'transaksi': conn.execute("SELECT COUNT(*) FROM transaksi").fetchone()[0],
        'jurnal': conn.execute("SELECT COUNT(*) FROM jurnal").fetchone()[0],
        'db_size': os.path.getsize(DB_PATH),
    }
    conn.close()
    return render_template('admin/backup.html', backups=backups, stats=stats, total_size=total_size)

@app.route('/admin/backup/create', methods=['POST'])
@admin_required
def admin_backup_create():
    fname = create_backup('manual')
    flash(f'Backup berhasil dibuat: {fname}', 'success')
    return redirect(url_for('admin_backup'))

@app.route('/admin/backup/download/<filename>')
@admin_required
def admin_backup_download(filename):
    if '..' in filename or '/' in filename or '\\' in filename:
        flash('Nama file tidak valid.', 'danger')
        return redirect(url_for('admin_backup'))
    fpath = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(fpath):
        flash('File backup tidak ditemukan.', 'danger')
        return redirect(url_for('admin_backup'))
    return send_file(fpath, download_name=filename, as_attachment=True)

@app.route('/admin/backup/delete/<filename>', methods=['POST'])
@admin_required
def admin_backup_delete(filename):
    if '..' in filename or '/' in filename or '\\' in filename:
        flash('Nama file tidak valid.', 'danger')
        return redirect(url_for('admin_backup'))
    fpath = os.path.join(BACKUP_DIR, filename)
    if os.path.exists(fpath):
        os.remove(fpath)
        flash(f'Backup {filename} dihapus.', 'warning')
    return redirect(url_for('admin_backup'))

@app.route('/admin/backup/restore/<filename>', methods=['POST'])
@admin_required
def admin_backup_restore(filename):
    if '..' in filename or '/' in filename or '\\' in filename:
        flash('Nama file tidak valid.', 'danger')
        return redirect(url_for('admin_backup'))
    fpath = os.path.join(BACKUP_DIR, filename)
    if not os.path.exists(fpath):
        flash('File backup tidak ditemukan.', 'danger')
        return redirect(url_for('admin_backup'))
    create_backup('pre_restore')
    src = sqlite3.connect(fpath)
    dst = sqlite3.connect(DB_PATH)
    src.backup(dst)
    dst.close()
    src.close()
    flash(f'Database berhasil di-restore dari {filename}. Backup sebelum restore juga dibuat.', 'success')
    return redirect(url_for('admin_backup'))

@app.route('/api/backup/auto', methods=['POST'])
def api_backup_auto():
    key = request.headers.get('X-Backup-Key', '')
    expected = os.environ.get('BACKUP_KEY') or app.secret_key
    if not key or key != expected:
        return jsonify(ok=False, msg='Unauthorized'), 401
    fname = create_backup('auto')
    cleanup = 0
    backups = sorted(glob_mod.glob(os.path.join(BACKUP_DIR, 'backup_auto_*.db')))
    while len(backups) > 12:
        os.remove(backups.pop(0))
        cleanup += 1
    return jsonify(ok=True, filename=fname, cleaned=cleanup)

# ── API read-only utk Agent Laporan Harian Pengelola ─────────────────────────
@app.route('/api/laporan-harian')
def api_laporan_harian():
    """Dipakai agent laporan harian (server-to-server). READ-ONLY.
    Auth: header X-Api-Key == env LAPORAN_API_KEY (fallback app.secret_key).
    Param: tanggal=YYYY-MM-DD (default hari ini).
    """
    key = request.headers.get('X-Api-Key', '')
    expected = os.environ.get('LAPORAN_API_KEY', '') or app.secret_key
    if not key or key != expected:
        return jsonify(error='unauthorized'), 401

    tgl = request.args.get('tanggal') or date.today().isoformat()
    conn = get_db()
    try:
        # Ambil semua user marketing aktif
        users = conn.execute(
            "SELECT id, nama FROM users WHERE role='marketing' AND aktif=1 ORDER BY nama"
        ).fetchall()

        # Ambil sum setoran ZIS/wakaf per user untuk tanggal tsb
        # Hanya dana ZIS/wakaf: zakat, infak (tidak terikat & terikat), wakaf
        DANA_ZIS = ('zakat', 'infak_tidak_terikat', 'infak_terikat', 'wakaf')
        placeholders = ','.join('?' * len(DANA_ZIS))
        rows_setor = conn.execute(f'''
            SELECT user_id, COALESCE(SUM(jumlah), 0) AS total_setoran
            FROM transaksi
            WHERE tanggal = ?
              AND jenis = 'masuk'
              AND jenis_dana IN ({placeholders})
            GROUP BY user_id
        ''', [tgl] + list(DANA_ZIS)).fetchall()

        setoran_map = {r['user_id']: float(r['total_setoran'] or 0) for r in rows_setor}

        petugas = []
        total_setoran = 0.0
        for u in users:
            setor = setoran_map.get(u['id'], 0.0)
            total_setoran += setor
            petugas.append({
                'nama': u['nama'],
                'setoran': setor,
                'donatur_baru': None
            })

        donatur_baru = conn.execute(
            "SELECT COUNT(*) c FROM donatur WHERE substr(COALESCE(created_at,''),1,10)=?",
            (tgl,)).fetchone()['c']

    finally:
        conn.close()

    return jsonify({
        'tanggal': tgl,
        'petugas': petugas,
        'total_setoran': total_setoran,
        'donatur_baru_lembaga': donatur_baru,
        # Belum terlacak di DB maal (perlu pencatatan terpisah / input manual):
        'tidak_tersedia': ['layanan_ambulan', 'penawaran_donasi',
                           'donatur_baru_per_petugas'],
    })

# ── Database Export / Import (Excel) ─────────────────────────────────────────

DB_EXPORT_TABLES = [
    {
        'name': 'donatur', 'label': 'Donatur',
        'query': "SELECT id, nama, no_hp, nik, alamat, jenis, sumber_infaq, area, lokasi_nama, lat, lng, aktif, aktif_infaq, program_id, created_at FROM donatur ORDER BY id",
        'headers': ['id','nama','no_hp','nik','alamat','jenis','sumber_infaq','area','lokasi_nama','lat','lng','aktif','aktif_infaq','program_id','created_at'],
        'color': '27AE60', 'import': True,
        'required': ['nama'],
    },
    {
        'name': 'penerima_manfaat', 'label': 'Penerima Manfaat',
        'query': "SELECT id, nama, nik, no_hp, alamat, asnaf, keterangan, aktif, created_at FROM penerima_manfaat ORDER BY id",
        'headers': ['id','nama','nik','no_hp','alamat','asnaf','keterangan','aktif','created_at'],
        'color': 'E74C3C', 'import': True,
        'required': ['nama'],
    },
    {
        'name': 'chart_of_accounts', 'label': 'Chart of Accounts',
        'query': "SELECT id, kode, nama, kelompok, jenis_dana, parent_kode, jenis_transaksi, aktif FROM chart_of_accounts ORDER BY kode",
        'headers': ['id','kode','nama','kelompok','jenis_dana','parent_kode','jenis_transaksi','aktif'],
        'color': '2980B9', 'import': True,
        'required': ['kode','nama','kelompok'],
    },
    {
        'name': 'users', 'label': 'Users',
        'query': "SELECT id, username, nama, role, no_hp, aktif FROM users ORDER BY id",
        'headers': ['id','username','nama','role','no_hp','aktif'],
        'color': '8E44AD', 'import': True,
        'required': ['username','nama','role'],
    },
    {
        'name': 'area', 'label': 'Area',
        'query': "SELECT id, nama, aktif FROM area ORDER BY nama",
        'headers': ['id','nama','aktif'],
        'color': 'E67E22', 'import': True,
        'required': ['nama'],
    },
    {
        'name': 'transaksi', 'label': 'Transaksi',
        'query': '''SELECT t.id, t.tanggal, t.jenis, t.jenis_dana, t.coa_id,
                     c.kode as coa_kode, c.nama as coa_nama,
                     t.donatur_id, d.nama as donatur_nama,
                     t.penerima_id, p.nama as penerima_nama,
                     t.jumlah, t.keterangan, t.user_id, u.nama as user_nama, t.created_at
                    FROM transaksi t
                    LEFT JOIN chart_of_accounts c ON t.coa_id=c.id
                    LEFT JOIN donatur d ON t.donatur_id=d.id
                    LEFT JOIN penerima_manfaat p ON t.penerima_id=p.id
                    LEFT JOIN users u ON t.user_id=u.id
                    ORDER BY t.tanggal DESC, t.id DESC''',
        'headers': ['id','tanggal','jenis','jenis_dana','coa_id','coa_kode','coa_nama',
                     'donatur_id','donatur_nama','penerima_id','penerima_nama',
                     'jumlah','keterangan','user_id','user_nama','created_at'],
        'color': '1A5276', 'import': False,
    },
    {
        'name': 'koleksi_bulanan', 'label': 'Koleksi Bulanan',
        'query': '''SELECT kb.id, kb.donatur_id, d.nama as donatur_nama, kb.bulan, kb.status,
                     kb.marketing_id, u.nama as marketing_nama,
                     kb.tanggal_koleksi, kb.jumlah, kb.jumlah_kunjungan,
                     kb.kunjungan_terakhir, kb.keterangan, kb.lat_kunjungan, kb.lng_kunjungan, kb.created_at
                    FROM koleksi_bulanan kb
                    JOIN donatur d ON kb.donatur_id=d.id
                    LEFT JOIN users u ON kb.marketing_id=u.id
                    ORDER BY kb.bulan DESC, d.nama''',
        'headers': ['id','donatur_id','donatur_nama','bulan','status',
                     'marketing_id','marketing_nama','tanggal_koleksi','jumlah',
                     'jumlah_kunjungan','kunjungan_terakhir','keterangan',
                     'lat_kunjungan','lng_kunjungan','created_at'],
        'color': 'D35400', 'import': False,
    },
    {
        'name': 'jurnal', 'label': 'Jurnal',
        'query': '''SELECT j.id, j.tanggal, j.no_bukti, j.keterangan,
                     j.debit_coa_id, cd.nama as debit_coa_nama,
                     j.kredit_coa_id, ck.nama as kredit_coa_nama,
                     j.jumlah, j.user_id, u.nama as user_nama, j.created_at
                    FROM jurnal j
                    LEFT JOIN chart_of_accounts cd ON j.debit_coa_id=cd.id
                    LEFT JOIN chart_of_accounts ck ON j.kredit_coa_id=ck.id
                    LEFT JOIN users u ON j.user_id=u.id
                    ORDER BY j.tanggal DESC, j.id DESC''',
        'headers': ['id','tanggal','no_bukti','keterangan','debit_coa_id','debit_coa_nama',
                     'kredit_coa_id','kredit_coa_nama','jumlah','user_id','user_nama','created_at'],
        'color': '566573', 'import': False,
    },
    {
        'name': 'instansi', 'label': 'Instansi',
        'query': "SELECT * FROM instansi",
        'headers': ['id','nama','nama_lembaga','alamat','telepon','email','website','ketua','bendahara','sekretaris','no_izin','updated_at'],
        'color': '34495E', 'import': True,
        'required': ['nama'],
    },
    {
        'name': 'target_bulanan', 'label': 'Target Bulanan',
        'query': '''SELECT tb.id, tb.bulan, tb.user_id, u.nama as user_nama,
                     tb.jenis, tb.target_nominal, tb.target_kegiatan, tb.created_at
                    FROM target_bulanan tb
                    LEFT JOIN users u ON tb.user_id=u.id
                    ORDER BY tb.bulan DESC, tb.jenis''',
        'headers': ['id','bulan','user_id','user_nama','jenis','target_nominal','target_kegiatan','created_at'],
        'color': '16A085', 'import': True,
        'required': ['bulan','jenis'],
    },
]

@app.route('/admin/db')
@admin_required
def admin_db():
    conn = get_db()
    counts = {}
    for tbl in DB_EXPORT_TABLES:
        try:
            counts[tbl['name']] = conn.execute(f"SELECT COUNT(*) FROM {tbl['name']}").fetchone()[0]
        except Exception:
            counts[tbl['name']] = 0
    conn.close()
    return render_template('admin/db_excel.html', tables=DB_EXPORT_TABLES, counts=counts)

@app.route('/admin/db/export-all')
@admin_required
def admin_db_export_all():
    conn = get_db()
    wb = Workbook()
    wb.remove(wb.active)
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    thin = Side(style='thin', color='DDDDDD')
    border = Border(bottom=thin)
    for tbl in DB_EXPORT_TABLES:
        ws = wb.create_sheet(title=tbl['label'][:31])
        fill = PatternFill('solid', fgColor=tbl['color'])
        for col, h in enumerate(tbl['headers'], 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center')
        try:
            rows = conn.execute(tbl['query']).fetchall()
        except Exception:
            rows = []
        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = border
        for col_idx in range(1, len(tbl['headers'])+1):
            max_len = len(tbl['headers'][col_idx-1])
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                val_len = len(str(row[0] or ''))
                if val_len > max_len:
                    max_len = val_len
            ws.column_dimensions[chr(64+col_idx) if col_idx <= 26 else 'A' + chr(64+col_idx-26)].width = min(max_len + 4, 40)
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = 'A2'
    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(buf, download_name=f'database_maal_{ts}.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/db/export/<table_name>')
@admin_required
def admin_db_export_table(table_name):
    tbl = next((t for t in DB_EXPORT_TABLES if t['name'] == table_name), None)
    if not tbl:
        flash('Tabel tidak ditemukan.', 'danger')
        return redirect(url_for('admin_db'))
    conn = get_db()
    wb = Workbook()
    ws = wb.active
    ws.title = tbl['label'][:31]
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    fill = PatternFill('solid', fgColor=tbl['color'])
    thin = Side(style='thin', color='DDDDDD')
    border = Border(bottom=thin)
    for col, h in enumerate(tbl['headers'], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    rows = conn.execute(tbl['query']).fetchall()
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
    for col_idx in range(1, len(tbl['headers'])+1):
        max_len = len(tbl['headers'][col_idx-1])
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            val_len = len(str(row[0] or ''))
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[chr(64+col_idx) if col_idx <= 26 else 'A' + chr(64+col_idx-26)].width = min(max_len + 4, 40)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'
    conn.close()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name=f'{table_name}.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/db/import', methods=['POST'])
@admin_required
def admin_db_import():
    f = request.files.get('file')
    if not f or not f.filename.endswith(('.xlsx', '.xls')):
        flash('Upload file Excel (.xlsx) yang valid.', 'danger')
        return redirect(url_for('admin_db'))
    try:
        wb = load_workbook(f, data_only=True)
    except Exception as e:
        flash(f'Gagal membaca file: {e}', 'danger')
        return redirect(url_for('admin_db'))

    create_backup('pre_import')
    conn = get_db()
    report = []

    importable = {t['name']: t for t in DB_EXPORT_TABLES if t.get('import')}
    IMPORT_COLS = {
        'donatur': ['nama','no_hp','nik','alamat','jenis','sumber_infaq','area','lokasi_nama','lat','lng','aktif','aktif_infaq','program_id'],
        'penerima_manfaat': ['nama','nik','no_hp','alamat','asnaf','keterangan','aktif'],
        'chart_of_accounts': ['kode','nama','kelompok','jenis_dana','parent_kode','jenis_transaksi','aktif'],
        'users': ['username','nama','role','no_hp','aktif'],
        'area': ['nama','aktif'],
        'instansi': ['nama','nama_lembaga','alamat','telepon','email','website','ketua','bendahara','sekretaris','no_izin'],
        'target_bulanan': ['bulan','user_id','jenis','target_nominal','target_kegiatan'],
    }

    for ws in wb.worksheets:
        sheet_label = ws.title.strip()
        tbl = next((t for t in DB_EXPORT_TABLES if t.get('import') and t['label'] == sheet_label), None)
        if not tbl:
            continue

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if len(rows) < 2:
            report.append(f"{sheet_label}: kosong, dilewati")
            continue

        header = [str(h).strip().lower() if h else '' for h in rows[0]]
        col_map = {h: i for i, h in enumerate(header) if h}
        tbl_name = tbl['name']
        allowed_cols = IMPORT_COLS.get(tbl_name, [])
        required = set(tbl.get('required', []))

        if not required.issubset(set(col_map.keys())):
            missing = required - set(col_map.keys())
            report.append(f"{sheet_label}: kolom wajib tidak ada ({', '.join(missing)})")
            continue

        has_id = 'id' in col_map
        updated = 0
        inserted = 0
        skipped = 0

        for row in rows[1:]:
            row_id = None
            if has_id:
                raw_id = row[col_map['id']]
                if raw_id is not None and str(raw_id).strip():
                    try:
                        row_id = int(float(str(raw_id)))
                    except (ValueError, TypeError):
                        row_id = None

            vals = {}
            for col_name in allowed_cols:
                if col_name in col_map:
                    v = row[col_map[col_name]]
                    vals[col_name] = v if v is not None else None

            req_ok = all(vals.get(r) not in (None, '', 'None', 'none') for r in required)
            if not req_ok:
                skipped += 1
                continue

            for c in ('aktif', 'aktif_infaq', 'program_id', 'target_nominal', 'target_kegiatan', 'user_id'):
                if c in vals and vals[c] is not None:
                    try:
                        vals[c] = int(float(str(vals[c])))
                    except (ValueError, TypeError):
                        vals[c] = None
            for c in ('lat', 'lng'):
                if c in vals and vals[c] is not None:
                    try:
                        vals[c] = float(str(vals[c]))
                    except (ValueError, TypeError):
                        vals[c] = None

            cols = list(vals.keys())
            values = [vals[c] for c in cols]

            if row_id:
                existing = conn.execute(f"SELECT id FROM {tbl_name} WHERE id=?", (row_id,)).fetchone()
                if existing:
                    set_clause = ', '.join(f'{c}=?' for c in cols)
                    conn.execute(f"UPDATE {tbl_name} SET {set_clause} WHERE id=?", values + [row_id])
                    updated += 1
                    continue

            placeholders = ','.join(['?'] * len(cols))
            col_names = ','.join(cols)
            try:
                conn.execute(f"INSERT INTO {tbl_name} ({col_names}) VALUES ({placeholders})", values)
                inserted += 1
            except Exception:
                skipped += 1

        report.append(f"{sheet_label}: {inserted} ditambahkan, {updated} diperbarui, {skipped} dilewati")

    conn.commit()
    conn.close()
    wb.close()
    flash('Import selesai. ' + ' | '.join(report), 'success' if report else 'warning')
    return redirect(url_for('admin_db'))

# ── Marketing Dashboard ───────────────────────────────────────────────────────

@app.route('/marketing')
@login_required
def marketing_dashboard():
    conn = get_db()
    bulan = date.today().strftime('%Y-%m')
    total_masuk_bulan = conn.execute(
        "SELECT COALESCE(SUM(jumlah),0) FROM transaksi WHERE jenis='masuk' AND strftime('%Y-%m',tanggal)=? AND user_id=?",
        (bulan, session['user_id'])
    ).fetchone()[0]
    koleksi_bulan = conn.execute(
        "SELECT COUNT(*) as jumlah, COALESCE(SUM(jumlah),0) as nominal FROM koleksi_bulanan "
        "WHERE bulan=? AND marketing_id=? AND status='terkumpul'",
        (bulan, session['user_id'])
    ).fetchone()
    transaksi_hari = conn.execute('''
        SELECT t.*, c.nama as coa_nama, c.jenis_dana, d.nama as donatur_nama
        FROM transaksi t
        LEFT JOIN chart_of_accounts c ON t.coa_id=c.id
        LEFT JOIN donatur d ON t.donatur_id=d.id
        WHERE t.user_id=? AND t.tanggal=? ORDER BY t.created_at DESC
    ''', (session['user_id'], date.today().isoformat())).fetchall()
    conn.close()
    return render_template('marketing/dashboard.html',
        total_masuk_bulan=total_masuk_bulan, koleksi_bulan=koleksi_bulan,
        transaksi_hari=transaksi_hari, hari_ini=date.today().strftime('%d %B %Y'))

@app.route('/marketing/saldo-program')
@login_required
def marketing_saldo_program():
    """Saldo akhir tiap program fundraising -- versi ringkas (kartu, read-only) dr
    laporan Saldo per Program admin, spy fundraiser bisa cek sisa saldo programnya."""
    bulan = get_tanggal_kerja()[:7]
    conn = get_db()
    groups, tidak_terikat, _ = _hitung_saldo_program(conn, bulan)
    data = _saldo_program_list(groups, tidak_terikat)
    conn.close()
    return render_template('marketing/saldo_program.html', data=data, bulan=bulan)

# ── Marketing Koleksi ─────────────────────────────────────────────────────────

@app.route('/marketing/koleksi')
@login_required
def marketing_koleksi():
    conn = get_db()
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    area  = request.args.get('area', '')
    q     = request.args.get('q', '').strip()
    query = '''
        SELECT kb.*, d.nama as donatur_nama, d.sumber_infaq, d.area,
               d.lokasi_nama, d.lat, d.lng, u.nama as marketing_nama
        FROM koleksi_bulanan kb
        JOIN donatur d ON kb.donatur_id=d.id
        LEFT JOIN users u ON kb.marketing_kunjungi_terakhir=u.id
        WHERE kb.bulan=? AND kb.status != 'terkumpul' '''
    params = [bulan]
    if area == '__none__':
        query += " AND (d.area IS NULL OR d.area='')"
    elif area:
        query += ' AND d.area=?'; params.append(area)
    if q:
        query += ' AND (d.nama LIKE ? OR d.lokasi_nama LIKE ?)'
        params += [f'%{q}%', f'%{q}%']
    query += ' ORDER BY d.area, d.sumber_infaq, d.nama'
    koleksi = conn.execute(query, params).fetchall()
    areas = conn.execute(
        "SELECT DISTINCT d.area FROM koleksi_bulanan kb JOIN donatur d ON kb.donatur_id=d.id "
        "WHERE kb.bulan=? AND kb.status!='terkumpul' AND d.area IS NOT NULL ORDER BY d.area",
        (bulan,)
    ).fetchall()
    stats = conn.execute(
        "SELECT COUNT(*) as total, SUM(CASE WHEN status='terkumpul' THEN 1 ELSE 0 END) as terkumpul "
        "FROM koleksi_bulanan WHERE bulan=?", (bulan,)
    ).fetchone()
    conn.close()
    return render_template('marketing/koleksi.html', koleksi=koleksi, areas=areas,
        bulan=bulan, area=area, q=q, stats=stats)

@app.route('/marketing/koleksi/<int:id>', methods=['GET'])
@login_required
def marketing_koleksi_form(id):
    conn = get_db()
    koleksi = conn.execute('''
        SELECT kb.*, d.nama as donatur_nama, d.sumber_infaq, d.area,
               d.lokasi_nama, d.lat, d.lng
        FROM koleksi_bulanan kb JOIN donatur d ON kb.donatur_id=d.id
        WHERE kb.id=?
    ''', (id,)).fetchone()
    if not koleksi:
        flash('Data koleksi tidak ditemukan.', 'danger')
        return redirect(url_for('marketing_koleksi'))
    if koleksi['status'] == 'terkumpul':
        flash('Kencleng/kotak ini sudah dikoleksi.', 'warning')
        return redirect(url_for('marketing_koleksi'))
    riwayat = conn.execute('''
        SELECT kb.*, u.nama as marketing_nama
        FROM koleksi_bulanan kb LEFT JOIN users u ON kb.marketing_id=u.id
        WHERE kb.donatur_id=? AND kb.status='terkumpul'
        ORDER BY kb.bulan DESC LIMIT 6
    ''', (koleksi['donatur_id'],)).fetchall()
    conn.close()
    return render_template('marketing/koleksi_catat.html', koleksi=koleksi,
        riwayat=riwayat, hari_ini=date.today().isoformat())

@app.route('/marketing/koleksi/<int:id>/catat', methods=['POST'])
@login_required
def marketing_koleksi_catat(id):
    conn  = get_db()
    aksi  = request.form.get('aksi')   # 'terkumpul' atau 'tidak_ada'
    today = date.today().isoformat()

    if aksi == 'terkumpul':
        try:
            jumlah = float((request.form.get('jumlah', '0') or '0').replace('.','').replace(',',''))
        except ValueError:
            conn.close(); flash('Jumlah tidak valid.', 'danger')
            return redirect(url_for('marketing_koleksi_form', id=id))
        kol = conn.execute("SELECT * FROM koleksi_bulanan WHERE id=?", (id,)).fetchone()
        if not kol:
            conn.close(); flash('Data tidak ditemukan.', 'danger')
            return redirect(url_for('marketing_koleksi'))

        rows = conn.execute(
            "UPDATE koleksi_bulanan SET status='terkumpul', marketing_id=?, "
            "tanggal_koleksi=?, jumlah=?, keterangan=? "
            "WHERE id=? AND status != 'terkumpul'",
            (session['user_id'], today, jumlah,
             request.form.get('keterangan',''), id)
        ).rowcount

        if rows == 0:
            # Cek siapa yang sudah koleksi duluan
            other = conn.execute(
                "SELECT u.nama, kb.tanggal_koleksi FROM koleksi_bulanan kb "
                "JOIN users u ON kb.marketing_id=u.id WHERE kb.id=?", (id,)
            ).fetchone()
            conn.close()
            msg = f"Sudah dikoleksi oleh {other['nama']} pada {other['tanggal_koleksi']}." if other else "Sudah dikoleksi."
            flash(msg, 'warning')
            return redirect(url_for('marketing_koleksi'))

        donatur = conn.execute("SELECT * FROM donatur WHERE id=?", (kol['donatur_id'],)).fetchone()
        trx_id = auto_transaksi_koleksi(conn, id, kol['donatur_id'], kol['bulan'],
                               donatur['sumber_infaq'], jumlah, today, session['user_id'])
        conn.commit(); conn.close()
        flash(f'Koleksi berhasil dicatat: {format_rupiah(jumlah)}', 'success')
        if jumlah > 0:
            return redirect(url_for('slip', transaksi_id=trx_id))

    elif aksi == 'tidak_ada':
        conn.execute(
            "UPDATE koleksi_bulanan SET status='tidak_ada', "
            "jumlah_kunjungan = jumlah_kunjungan + 1, "
            "kunjungan_terakhir=?, marketing_kunjungi_terakhir=?, keterangan=? "
            "WHERE id=?",
            (today, session['user_id'], request.form.get('keterangan',''), id)
        )
        conn.commit(); conn.close()
        flash('Kunjungan dicatat. Coba lagi lain waktu.', 'info')

    return redirect(url_for('marketing_koleksi'))

# ── Marketing Donatur Detail ──────────────────────────────────────────────────

@app.route('/marketing/donatur/<int:id>')
@login_required
def marketing_donatur_detail(id):
    conn = get_db()
    donatur = conn.execute("SELECT * FROM donatur WHERE id=?", (id,)).fetchone()
    if not donatur:
        conn.close(); flash('Donatur tidak ditemukan.', 'danger')
        return redirect(url_for('marketing_koleksi'))
    riwayat = conn.execute('''
        SELECT kb.*, u.nama as marketing_nama
        FROM koleksi_bulanan kb LEFT JOIN users u ON kb.marketing_id=u.id
        WHERE kb.donatur_id=? ORDER BY kb.bulan DESC
    ''', (id,)).fetchall()
    areas = conn.execute("SELECT nama AS area FROM area WHERE aktif=1 ORDER BY nama").fetchall()
    conn.close()
    return render_template('marketing/donatur_detail.html', donatur=donatur,
                           riwayat=riwayat, areas=areas)

# ── Marketing Peta ────────────────────────────────────────────────────────────

@app.route('/marketing/peta')
@login_required
def marketing_peta():
    conn = get_db()
    bulan = date.today().strftime('%Y-%m')
    titik = conn.execute('''
        SELECT kb.id, kb.status, kb.jumlah_kunjungan,
               d.nama, d.sumber_infaq, d.area, d.lokasi_nama, d.lat, d.lng
        FROM koleksi_bulanan kb JOIN donatur d ON kb.donatur_id=d.id
        WHERE kb.bulan=? AND kb.status != 'terkumpul'
              AND d.lat IS NOT NULL AND d.lng IS NOT NULL
        ORDER BY d.area, d.nama
    ''', (bulan,)).fetchall()
    titik_json = json.dumps([dict(t) for t in titik])
    desa_points = conn.execute('''
        SELECT nama, desa, lat, lng FROM donatur
        WHERE aktif=1 AND lat IS NOT NULL AND lng IS NOT NULL
              AND desa IS NOT NULL AND TRIM(desa) <> ''
    ''').fetchall()
    desa_json = json.dumps([dict(d) for d in desa_points])
    conn.close()
    return render_template('marketing/peta.html', titik_json=titik_json,
                           desa_json=desa_json, bulan=bulan)

# ── Marketing Donatur ─────────────────────────────────────────────────────────

@app.route('/marketing/donatur')
@login_required
def marketing_donatur_list():
    conn = get_db()
    q = request.args.get('q', '').strip()
    query = "SELECT * FROM donatur WHERE aktif=1"
    params = []
    if q:
        query += " AND (nama LIKE ? OR area LIKE ? OR no_hp LIKE ?)"
        params += [f'%{q}%'] * 3
    query += " ORDER BY nama"
    donatur = conn.execute(query, params).fetchall()
    areas = conn.execute("SELECT nama AS area FROM area WHERE aktif=1 ORDER BY nama").fetchall()
    conn.close()
    return render_template('marketing/donatur.html', donatur=donatur, areas=areas, q=q)

@app.route('/marketing/donatur/tambah', methods=['POST'])
@login_required
def marketing_donatur_tambah():
    data = request.form
    lat = lng = None
    gmaps = data.get('gmaps_url', '').strip()
    if gmaps:
        lat, lng = parse_gmaps_url(gmaps)
    if lat is None and data.get('lat'):
        try: lat = float(data['lat']); lng = float(data['lng'])
        except (TypeError, ValueError): lat = lng = None
    conn = get_db()
    sumber = data.get('sumber_infaq', 'tunai')
    nama = (data.get('nama') or '').strip()
    if not nama:
        conn.close(); flash('Nama donatur wajib diisi.', 'danger')
        return redirect(url_for('marketing_donatur_list'))
    cur = conn.execute("""INSERT INTO donatur
        (nama,no_hp,alamat,sumber_infaq,area,lokasi_nama,lat,lng,aktif_infaq)
        VALUES (?,?,?,?,?,?,?,?,?)""",
        (nama, data.get('no_hp', '').strip(), data.get('alamat', '').strip(),
         sumber,
         data.get('area', '').strip(), data.get('lokasi_nama', '').strip(),
         lat, lng, 1 if data.get('aktif_infaq') else 0))
    auto_koleksi_donatur_baru(conn, cur.lastrowid, sumber)
    conn.commit(); conn.close()
    flash(f'Donatur "{nama}" berhasil ditambahkan.', 'success')
    return redirect(url_for('marketing_donatur_list'))

@app.route('/marketing/donatur/edit/<int:id>', methods=['POST'])
@login_required
def marketing_donatur_edit(id):
    data = request.form
    conn = get_db()
    existing = conn.execute("SELECT * FROM donatur WHERE id=?", (id,)).fetchone()
    if not existing:
        conn.close(); flash('Donatur tidak ditemukan.', 'danger')
        return redirect(url_for('marketing_donatur_list'))
    nama = (data.get('nama') or '').strip()
    if not nama:
        conn.close(); flash('Nama donatur wajib diisi.', 'danger')
        return redirect(url_for('marketing_donatur_detail', id=id))
    lat = lng = None
    gmaps = data.get('gmaps_url', '').strip()
    if gmaps:
        lat, lng = parse_gmaps_url(gmaps)
    if lat is None and data.get('lat'):
        try: lat = float(data['lat']); lng = float(data['lng'])
        except (TypeError, ValueError): lat = lng = None
    if lat is None:
        lat, lng = existing['lat'], existing['lng']
    sumber = data.get('sumber_infaq') or existing['sumber_infaq']
    conn.execute("""UPDATE donatur SET
        nama=?, no_hp=?, alamat=?, sumber_infaq=?, area=?, lokasi_nama=?,
        lat=?, lng=?, aktif_infaq=?
        WHERE id=?""",
        (nama, data.get('no_hp', '').strip(), data.get('alamat', '').strip(),
         sumber, data.get('area', '').strip(), data.get('lokasi_nama', '').strip(),
         lat, lng, 1 if data.get('aktif_infaq') else 0, id))
    auto_koleksi_donatur_baru(conn, id, sumber)
    conn.commit(); conn.close()
    flash('Data donatur berhasil diperbarui.', 'success')
    return redirect(url_for('marketing_donatur_detail', id=id))

# ── Marketing Transaksi Tunai ─────────────────────────────────────────────────

@app.route('/marketing/catat', methods=['GET', 'POST'])
@login_required
def marketing_catat():
    conn = get_db()
    if request.method == 'POST':
        data = request.form
        if data.get('jenis') not in ('masuk', 'keluar'):
            conn.close(); flash('Jenis transaksi tidak valid.', 'danger')
            return redirect(url_for('marketing_catat'))
        jumlah = parse_jumlah(data.get('jumlah'))
        if jumlah is None:
            conn.close(); flash('Jumlah tidak valid — harus angka lebih dari 0.', 'danger')
            return redirect(url_for('marketing_catat'))
        jumlah_mustahik = None
        if data.get('jumlah_mustahik'):
            try: jumlah_mustahik = int(data['jumlah_mustahik'])
            except ValueError: jumlah_mustahik = None
        trx_id, dup = insert_transaksi(conn, data['tanggal'], data['jenis'],
            data.get('coa_id') or None, data.get('donatur_id') or None,
            data.get('penerima_id') or None, jumlah,
            data.get('keterangan',''), session['user_id'], data.get('client_uuid'),
            nama_kegiatan=data.get('nama_kegiatan') or None,
            lokasi=data.get('lokasi') or None,
            jumlah_mustahik=jumlah_mustahik)
        conn.commit(); conn.close()
        if dup:
            flash('Transaksi ini sudah tercatat sebelumnya — tidak dicatat ganda.', 'warning')
        else:
            flash('Transaksi berhasil dicatat!', 'success')
        if data['jenis'] == 'masuk':
            return redirect(url_for('slip', transaksi_id=trx_id))
        return redirect(url_for('marketing_dashboard'))
    coa_list      = conn.execute("SELECT * FROM chart_of_accounts WHERE jenis_transaksi IS NOT NULL AND aktif=1 ORDER BY kode").fetchall()
    coa_parents   = conn.execute("SELECT kode, nama FROM chart_of_accounts WHERE parent_kode IS NOT NULL AND aktif=1 ORDER BY kode").fetchall()
    donatur_list  = conn.execute("SELECT id, nama, area FROM donatur WHERE aktif=1 ORDER BY nama").fetchall()
    penerima_list = conn.execute("SELECT * FROM penerima_manfaat WHERE aktif=1 ORDER BY nama").fetchall()
    conn.close()
    return render_template('marketing/catat.html', coa_list=coa_list,
        coa_parents=coa_parents,
        donatur_list=donatur_list, penerima_list=penerima_list,
        hari_ini=date.today().isoformat())

@app.route('/marketing/riwayat')
@login_required
def marketing_riwayat():
    conn = get_db()
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    transaksi = conn.execute('''
        SELECT t.*, c.nama as coa_nama, c.jenis_dana, d.nama as donatur_nama
        FROM transaksi t
        LEFT JOIN chart_of_accounts c ON t.coa_id=c.id
        LEFT JOIN donatur d ON t.donatur_id=d.id
        WHERE t.user_id=? AND strftime('%Y-%m',t.tanggal)=?
        ORDER BY t.tanggal DESC, t.created_at DESC
    ''', (session['user_id'], bulan)).fetchall()
    total = sum(r['jumlah'] for r in transaksi if r['jenis']=='masuk')
    conn.close()
    return render_template('marketing/riwayat.html', transaksi=transaksi, bulan=bulan, total=total)

# ── PWA Service Worker ────────────────────────────────────────────────────────

@app.route('/marketing/sw.js')
def marketing_sw():
    return app.send_static_file('sw.js'), 200, {
        'Content-Type': 'application/javascript',
        'Service-Worker-Allowed': '/marketing/'
    }

# ── Marketing API (JSON) ─────────────────────────────────────────────────────

def api_login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated

@app.route('/api/marketing/dashboard')
@api_login_required
def api_marketing_dashboard():
    conn = get_db()
    bulan = date.today().strftime('%Y-%m')
    total_masuk_bulan = conn.execute(
        "SELECT COALESCE(SUM(jumlah),0) FROM transaksi WHERE jenis='masuk' AND strftime('%Y-%m',tanggal)=? AND user_id=?",
        (bulan, session['user_id'])
    ).fetchone()[0]
    koleksi_bulan = conn.execute(
        "SELECT COUNT(*) as jumlah, COALESCE(SUM(jumlah),0) as nominal FROM koleksi_bulanan "
        "WHERE bulan=? AND marketing_id=? AND status='terkumpul'",
        (bulan, session['user_id'])
    ).fetchone()
    transaksi_hari = conn.execute('''
        SELECT t.id, t.tanggal, t.jenis, t.jumlah, t.keterangan,
               c.nama as coa_nama, c.jenis_dana, d.nama as donatur_nama
        FROM transaksi t
        LEFT JOIN chart_of_accounts c ON t.coa_id=c.id
        LEFT JOIN donatur d ON t.donatur_id=d.id
        WHERE t.user_id=? AND t.tanggal=? ORDER BY t.created_at DESC
    ''', (session['user_id'], date.today().isoformat())).fetchall()
    conn.close()
    return jsonify({
        'bulan': bulan,
        'total_masuk_bulan': total_masuk_bulan,
        'koleksi_jumlah': koleksi_bulan['jumlah'],
        'koleksi_nominal': koleksi_bulan['nominal'],
        'transaksi_hari': [dict(r) for r in transaksi_hari],
        'hari_ini': date.today().isoformat(),
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/marketing/koleksi')
@api_login_required
def api_marketing_koleksi():
    conn = get_db()
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    area  = request.args.get('area', '')
    query = '''
        SELECT kb.id, kb.donatur_id, kb.bulan, kb.status, kb.jumlah,
               kb.jumlah_kunjungan, kb.kunjungan_terakhir, kb.keterangan,
               d.nama as donatur_nama, d.sumber_infaq, d.area,
               d.lokasi_nama, d.lat, d.lng, u.nama as marketing_nama
        FROM koleksi_bulanan kb
        JOIN donatur d ON kb.donatur_id=d.id
        LEFT JOIN users u ON kb.marketing_kunjungi_terakhir=u.id
        WHERE kb.bulan=? AND kb.status != 'terkumpul' '''
    params = [bulan]
    if area:
        query += ' AND d.area=?'; params.append(area)
    query += ' ORDER BY d.area, d.sumber_infaq, d.nama'
    koleksi = conn.execute(query, params).fetchall()
    areas = conn.execute(
        "SELECT DISTINCT d.area FROM koleksi_bulanan kb JOIN donatur d ON kb.donatur_id=d.id "
        "WHERE kb.bulan=? AND kb.status!='terkumpul' AND d.area IS NOT NULL ORDER BY d.area",
        (bulan,)
    ).fetchall()
    stats = conn.execute(
        "SELECT COUNT(*) as total, SUM(CASE WHEN status='terkumpul' THEN 1 ELSE 0 END) as terkumpul "
        "FROM koleksi_bulanan WHERE bulan=?", (bulan,)
    ).fetchone()
    conn.close()
    return jsonify({
        'koleksi': [dict(r) for r in koleksi],
        'areas': [r['area'] for r in areas],
        'stats': dict(stats),
        'bulan': bulan,
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/marketing/koleksi/<int:id>')
@api_login_required
def api_marketing_koleksi_detail(id):
    conn = get_db()
    koleksi = conn.execute('''
        SELECT kb.*, d.nama as donatur_nama, d.sumber_infaq, d.area,
               d.lokasi_nama, d.lat, d.lng
        FROM koleksi_bulanan kb JOIN donatur d ON kb.donatur_id=d.id
        WHERE kb.id=?
    ''', (id,)).fetchone()
    if not koleksi:
        conn.close()
        return jsonify({'error': 'not_found'}), 404
    riwayat = conn.execute('''
        SELECT kb.bulan, kb.status, kb.jumlah, kb.tanggal_koleksi, u.nama as marketing_nama
        FROM koleksi_bulanan kb LEFT JOIN users u ON kb.marketing_id=u.id
        WHERE kb.donatur_id=? AND kb.status='terkumpul'
        ORDER BY kb.bulan DESC LIMIT 6
    ''', (koleksi['donatur_id'],)).fetchall()
    conn.close()
    return jsonify({
        'koleksi': dict(koleksi),
        'riwayat': [dict(r) for r in riwayat]
    })

@app.route('/api/marketing/koleksi/<int:id>/catat', methods=['POST'])
@api_login_required
def api_marketing_koleksi_catat(id):
    data  = request.get_json(silent=True) or {}
    aksi  = data.get('aksi')
    today = date.today().isoformat()
    conn  = get_db()

    lat_k = data.get('lat_kunjungan')
    lng_k = data.get('lng_kunjungan')

    if aksi == 'terkumpul':
        jumlah = float(data.get('jumlah', 0) or 0)
        kol = conn.execute("SELECT * FROM koleksi_bulanan WHERE id=?", (id,)).fetchone()
        if not kol:
            conn.close()
            return jsonify({'status': 'error', 'message': 'Data tidak ditemukan.'}), 404

        rows = conn.execute(
            "UPDATE koleksi_bulanan SET status='terkumpul', marketing_id=?, "
            "tanggal_koleksi=?, jumlah=?, keterangan=?, lat_kunjungan=?, lng_kunjungan=? "
            "WHERE id=? AND status != 'terkumpul'",
            (session['user_id'], today, jumlah, data.get('keterangan',''), lat_k, lng_k, id)
        ).rowcount

        if rows == 0:
            other = conn.execute(
                "SELECT u.nama, kb.tanggal_koleksi FROM koleksi_bulanan kb "
                "JOIN users u ON kb.marketing_id=u.id WHERE kb.id=?", (id,)
            ).fetchone()
            conn.close()
            msg = f"Sudah dikoleksi oleh {other['nama']} pada {other['tanggal_koleksi']}." if other else "Sudah dikoleksi."
            return jsonify({'status': 'conflict', 'message': msg}), 409

        donatur = conn.execute("SELECT * FROM donatur WHERE id=?", (kol['donatur_id'],)).fetchone()
        trx_id = auto_transaksi_koleksi(conn, id, kol['donatur_id'], kol['bulan'],
                                        donatur['sumber_infaq'], jumlah, today, session['user_id'])
        conn.commit(); conn.close()
        return jsonify({'status': 'ok', 'message': f'Koleksi berhasil: {format_rupiah(jumlah)}', 'transaksi_id': trx_id})

    elif aksi == 'tidak_ada':
        conn.execute(
            "UPDATE koleksi_bulanan SET status='tidak_ada', "
            "jumlah_kunjungan = jumlah_kunjungan + 1, "
            "kunjungan_terakhir=?, marketing_kunjungi_terakhir=?, keterangan=?, "
            "lat_kunjungan=?, lng_kunjungan=? "
            "WHERE id=?",
            (today, session['user_id'], data.get('keterangan',''), lat_k, lng_k, id)
        )
        conn.commit(); conn.close()
        return jsonify({'status': 'ok', 'message': 'Kunjungan dicatat.'})

    conn.close()
    return jsonify({'status': 'error', 'message': 'Aksi tidak valid.'}), 400

@app.route('/api/marketing/transaksi', methods=['POST'])
@api_login_required
def api_marketing_transaksi():
    data = request.get_json(silent=True) or {}
    if data.get('jenis') not in ('masuk', 'keluar'):
        return jsonify({'status': 'error', 'message': 'Jenis transaksi tidak valid.'}), 400
    jumlah = parse_jumlah(data.get('jumlah'))
    if jumlah is None:
        return jsonify({'status': 'error', 'message': 'Jumlah tidak valid — harus angka lebih dari 0.'}), 400
    jumlah_mustahik = None
    if data.get('jumlah_mustahik'):
        try: jumlah_mustahik = int(data['jumlah_mustahik'])
        except (TypeError, ValueError): jumlah_mustahik = None
    conn = get_db()
    trx_id, dup = insert_transaksi(conn,
        data.get('tanggal') or date.today().isoformat(),
        data['jenis'], data.get('coa_id') or None,
        data.get('donatur_id') or None, data.get('penerima_id') or None,
        jumlah, data.get('keterangan',''), session['user_id'],
        data.get('client_uuid'),
        nama_kegiatan=data.get('nama_kegiatan') or None,
        lokasi=data.get('lokasi') or None,
        jumlah_mustahik=jumlah_mustahik)
    conn.commit(); conn.close()
    msg = ('Transaksi ini sudah tercatat sebelumnya — tidak dicatat ganda.'
           if dup else 'Transaksi berhasil dicatat.')
    return jsonify({'status': 'ok', 'message': msg, 'transaksi_id': trx_id, 'duplicate': dup})

@app.route('/api/marketing/sync', methods=['POST'])
@api_login_required
def api_marketing_sync():
    items = (request.get_json(silent=True) or {}).get('items', [])
    results = []
    # Simpan identitas user SEBELUM masuk test_request_context (context baru = session kosong)
    cur_user = {'user_id': session.get('user_id'), 'role': session.get('role'), 'nama': session.get('nama')}
    for item in items:
        try:
            item_type = item.get('type')
            body = item.get('body', {})
            if item_type == 'koleksi_catat':
                with app.test_request_context(
                    f"/api/marketing/koleksi/{item['koleksi_id']}/catat",
                    method='POST', json=body
                ):
                    session.update(cur_user)
                    resp = api_marketing_koleksi_catat(item['koleksi_id'])
                    if isinstance(resp, tuple):
                        results.append({'id': item.get('id'), 'status': 'ok' if resp[1] < 400 else 'error',
                                        'response': resp[0].get_json()})
                    else:
                        results.append({'id': item.get('id'), 'status': 'ok', 'response': resp.get_json()})
            elif item_type == 'transaksi':
                jumlah = parse_jumlah(body.get('jumlah'))
                if body.get('jenis') not in ('masuk', 'keluar') or jumlah is None:
                    results.append({'id': item.get('id'), 'status': 'error',
                                    'message': 'Data transaksi tidak valid.'})
                    continue
                jumlah_mustahik = None
                if body.get('jumlah_mustahik'):
                    try: jumlah_mustahik = int(body['jumlah_mustahik'])
                    except (TypeError, ValueError): jumlah_mustahik = None
                conn = get_db()
                trx_id, dup = insert_transaksi(conn,
                    body.get('tanggal') or date.today().isoformat(),
                    body['jenis'], body.get('coa_id') or None,
                    body.get('donatur_id') or None, body.get('penerima_id') or None,
                    jumlah, body.get('keterangan',''), session['user_id'],
                    body.get('client_uuid'),
                    nama_kegiatan=body.get('nama_kegiatan') or None,
                    lokasi=body.get('lokasi') or None,
                    jumlah_mustahik=jumlah_mustahik)
                conn.commit(); conn.close()
                results.append({'id': item.get('id'), 'status': 'ok',
                                'transaksi_id': trx_id, 'duplicate': dup})
            else:
                results.append({'id': item.get('id'), 'status': 'error', 'message': 'Unknown type'})
        except Exception as ex:
            results.append({'id': item.get('id'), 'status': 'error', 'message': str(ex)})
    return jsonify({'results': results, 'synced': len([r for r in results if r['status']=='ok']), 'total': len(items)})

@app.route('/api/marketing/coa')
@api_login_required
def api_marketing_coa():
    conn = get_db()
    coa_list = conn.execute(
        "SELECT id, kode, nama, kelompok, jenis_dana, jenis_transaksi, parent_kode "
        "FROM chart_of_accounts WHERE aktif=1 ORDER BY kode"
    ).fetchall()
    conn.close()
    return jsonify({'coa': [dict(r) for r in coa_list]})

@app.route('/api/marketing/donatur')
@api_login_required
def api_marketing_donatur():
    conn = get_db()
    donatur = conn.execute(
        "SELECT id, nama, area, sumber_infaq FROM donatur WHERE aktif=1 ORDER BY nama"
    ).fetchall()
    penerima = conn.execute(
        "SELECT id, nama, asnaf FROM penerima_manfaat WHERE aktif=1 ORDER BY nama"
    ).fetchall()
    conn.close()
    return jsonify({'donatur': [dict(r) for r in donatur], 'penerima': [dict(r) for r in penerima]})

# ── Push Notification API ────────────────────────────────────────────────────

@app.route('/api/push/vapid-key')
def api_push_vapid_key():
    return jsonify({'publicKey': VAPID_PUBLIC_KEY})

@app.route('/api/push/subscribe', methods=['POST'])
@api_login_required
def api_push_subscribe():
    data = request.get_json(silent=True) or {}
    sub_json = json.dumps(data.get('subscription', {}))
    if not sub_json or sub_json == '{}':
        return jsonify({'error': 'No subscription data'}), 400
    conn = get_db()
    conn.execute(
        "INSERT OR IGNORE INTO push_subscriptions (user_id, subscription_json) VALUES (?, ?)",
        (session['user_id'], sub_json))
    conn.commit(); conn.close()
    return jsonify({'status': 'ok'})

@app.route('/api/push/unsubscribe', methods=['POST'])
@api_login_required
def api_push_unsubscribe():
    data = request.get_json(silent=True) or {}
    sub_json = json.dumps(data.get('subscription', {}))
    conn = get_db()
    conn.execute(
        "DELETE FROM push_subscriptions WHERE user_id=? AND subscription_json=?",
        (session['user_id'], sub_json))
    conn.commit(); conn.close()
    return jsonify({'status': 'ok'})

def send_push_to_user(user_id, title, body, url='/marketing'):
    if not VAPID_PRIVATE_KEY:
        return
    try:
        from pywebpush import webpush, WebPushException
    except ImportError:
        return
    conn = get_db()
    subs = conn.execute(
        "SELECT subscription_json FROM push_subscriptions WHERE user_id=?", (user_id,)
    ).fetchall()
    conn.close()
    payload = json.dumps({'title': title, 'body': body, 'url': url})
    for s in subs:
        try:
            webpush(
                subscription_info=json.loads(s['subscription_json']),
                data=payload,
                vapid_private_key=VAPID_PRIVATE_KEY,
                vapid_claims={'sub': f'mailto:{VAPID_CLAIM_EMAIL}'}
            )
        except Exception:
            pass

def send_push_to_all_marketing(title, body, url='/marketing'):
    if not VAPID_PRIVATE_KEY:
        return
    conn = get_db()
    users = conn.execute("SELECT id FROM users WHERE role='marketing' AND aktif=1").fetchall()
    conn.close()
    for u in users:
        send_push_to_user(u['id'], title, body, url)

# ── Target Bulanan API ───────────────────────────────────────────────────────

@app.route('/api/marketing/target')
@api_login_required
def api_marketing_target():
    conn = get_db()
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    uid = session['user_id']

    targets = conn.execute(
        "SELECT * FROM target_bulanan WHERE bulan=? AND (user_id=? OR user_id IS NULL) ORDER BY user_id DESC",
        (bulan, uid)
    ).fetchall()

    result = {}
    for jenis in ('fundraising', 'pentasharufan'):
        t = next((dict(r) for r in targets if r['jenis'] == jenis), None)
        if not t:
            result[jenis] = None
            continue

        if jenis == 'fundraising':
            realisasi = conn.execute(
                "SELECT COALESCE(SUM(jumlah),0) FROM transaksi "
                "WHERE jenis='masuk' AND strftime('%Y-%m',tanggal)=? AND user_id=?",
                (bulan, uid)
            ).fetchone()[0]
            result[jenis] = {
                'target_nominal': t['target_nominal'],
                'realisasi_nominal': realisasi,
                'persen': round(realisasi / t['target_nominal'] * 100, 1) if t['target_nominal'] > 0 else 0
            }
        else:
            realisasi_nominal = conn.execute(
                "SELECT COALESCE(SUM(jumlah),0) FROM transaksi "
                "WHERE jenis='keluar' AND strftime('%Y-%m',tanggal)=? AND user_id=?",
                (bulan, uid)
            ).fetchone()[0]
            realisasi_kegiatan = conn.execute(
                "SELECT COUNT(DISTINCT tanggal || coa_id) FROM transaksi "
                "WHERE jenis='keluar' AND strftime('%Y-%m',tanggal)=? AND user_id=?",
                (bulan, uid)
            ).fetchone()[0]
            result[jenis] = {
                'target_nominal': t['target_nominal'],
                'target_kegiatan': t['target_kegiatan'],
                'realisasi_nominal': realisasi_nominal,
                'realisasi_kegiatan': realisasi_kegiatan,
                'persen_nominal': round(realisasi_nominal / t['target_nominal'] * 100, 1) if t['target_nominal'] > 0 else 0,
                'persen_kegiatan': round(realisasi_kegiatan / t['target_kegiatan'] * 100, 1) if t['target_kegiatan'] > 0 else 0
            }

    conn.close()
    return jsonify({'bulan': bulan, 'target': result})

@app.route('/api/marketing/rkt')
@api_login_required
def api_marketing_rkt():
    """Ringkasan progres Rencana Kerja Tahunan (RKT) tahun berjalan -- dipanggil dashboard
    marketing tiap dibuka, spy RKT org (bukan cuma target pribadi bulanan) selalu jadi
    rujukan tiap bulan tanpa perlu setup apa pun tiap bulannya."""
    tahun = date.today().strftime('%Y')
    conn = get_db()
    ada_rkt = conn.execute("SELECT 1 FROM renstra_target WHERE tahun=? LIMIT 1", (tahun,)).fetchone() is not None
    if not ada_rkt:
        conn.close()
        return jsonify(ok=True, ada_rkt=False)
    totals = _renstra_totals(conn, tahun)
    conn.close()
    return jsonify(ok=True, ada_rkt=True, tahun=tahun,
        bulan_berjalan=int(date.today().strftime('%m')), **totals)

@app.route('/admin/target', methods=['GET', 'POST'])
@admin_required
def admin_target():
    conn = get_db()
    if request.method == 'POST':
        data = request.form
        bulan = data['bulan']
        user_id = int(data['user_id']) if data.get('user_id') else None
        jenis = data['jenis']
        nominal = int(data.get('target_nominal', 0) or 0)
        kegiatan = int(data.get('target_kegiatan', 0) or 0)
        conn.execute(
            "INSERT INTO target_bulanan (bulan, user_id, jenis, target_nominal, target_kegiatan) "
            "VALUES (?,?,?,?,?) ON CONFLICT(bulan, user_id, jenis) DO UPDATE SET "
            "target_nominal=excluded.target_nominal, target_kegiatan=excluded.target_kegiatan",
            (bulan, user_id, jenis, nominal, kegiatan))
        conn.commit()
        flash('Target berhasil disimpan.', 'success')
        return redirect(url_for('admin_target'))

    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    targets = conn.execute(
        "SELECT t.*, u.nama as user_nama FROM target_bulanan t "
        "LEFT JOIN users u ON t.user_id=u.id WHERE t.bulan=? ORDER BY t.jenis, u.nama",
        (bulan,)
    ).fetchall()
    marketing_users = conn.execute(
        "SELECT id, nama FROM users WHERE role='marketing' AND aktif=1 ORDER BY nama"
    ).fetchall()

    # ── Capaian real-time per marketing (rumus sama dgn /api/marketing/target) ──
    # Target personal menang atas target global (user_id NULL) — sama seperti
    # ORDER BY user_id DESC di api_marketing_target.
    tmap = {(t['user_id'], t['jenis']): t for t in targets}
    masuk_map = {r['user_id']: r for r in conn.execute("""
        SELECT user_id, COALESCE(SUM(jumlah),0) AS nominal, COUNT(*) AS n
        FROM transaksi WHERE jenis='masuk' AND strftime('%Y-%m',tanggal)=?
        GROUP BY user_id""", (bulan,)).fetchall()}
    keluar_map = {r['user_id']: r for r in conn.execute("""
        SELECT user_id, COALESCE(SUM(jumlah),0) AS nominal,
               COUNT(DISTINCT tanggal || coa_id) AS kegiatan
        FROM transaksi WHERE jenis='keluar' AND strftime('%Y-%m',tanggal)=?
        GROUP BY user_id""", (bulan,)).fetchall()}
    koleksi_map = {r['marketing_id']: r for r in conn.execute("""
        SELECT marketing_id, COUNT(*) AS n, COALESCE(SUM(jumlah),0) AS nominal
        FROM koleksi_bulanan WHERE bulan=? AND status='terkumpul'
        GROUP BY marketing_id""", (bulan,)).fetchall()}
    conn.close()

    def _persen(real, tgt):
        return round(real / tgt * 100, 1) if tgt else None

    capaian = []
    for u in marketing_users:
        tf = tmap.get((u['id'], 'fundraising'))   or tmap.get((None, 'fundraising'))
        tp = tmap.get((u['id'], 'pentasharufan')) or tmap.get((None, 'pentasharufan'))
        m, k, kol = masuk_map.get(u['id']), keluar_map.get(u['id']), koleksi_map.get(u['id'])
        row = {
            'nama': u['nama'],
            'fund_real':   m['nominal'] if m else 0,
            'fund_n':      m['n'] if m else 0,
            'fund_target': (tf['target_nominal'] if tf else 0) or 0,
            'pent_real':     k['nominal'] if k else 0,
            'pent_kegiatan': k['kegiatan'] if k else 0,
            'pent_target':          (tp['target_nominal'] if tp else 0) or 0,
            'pent_target_kegiatan': (tp['target_kegiatan'] if tp else 0) or 0,
            'koleksi_n':       kol['n'] if kol else 0,
            'koleksi_nominal': kol['nominal'] if kol else 0,
        }
        row['fund_persen']          = _persen(row['fund_real'], row['fund_target'])
        row['pent_persen']          = _persen(row['pent_real'], row['pent_target'])
        row['pent_persen_kegiatan'] = _persen(row['pent_kegiatan'], row['pent_target_kegiatan'])
        capaian.append(row)
    capaian.sort(key=lambda r: r['fund_real'], reverse=True)

    total = {
        'fund_real':     sum(r['fund_real'] for r in capaian),
        'fund_n':        sum(r['fund_n'] for r in capaian),
        'fund_target':   sum(r['fund_target'] for r in capaian),
        'pent_real':     sum(r['pent_real'] for r in capaian),
        'pent_kegiatan': sum(r['pent_kegiatan'] for r in capaian),
        'koleksi_n':     sum(r['koleksi_n'] for r in capaian),
    }
    total['fund_persen'] = _persen(total['fund_real'], total['fund_target'])

    return render_template('admin/target.html',
        targets=targets, marketing_users=marketing_users, bulan=bulan,
        capaian=capaian, total=total)

@app.route('/admin/target/delete/<int:id>', methods=['POST'])
@admin_required
def admin_target_delete(id):
    conn = get_db()
    conn.execute("DELETE FROM target_bulanan WHERE id=?", (id,))
    conn.commit(); conn.close()
    flash('Target dihapus.', 'success')
    return redirect(url_for('admin_target'))

# ── Slip Penerimaan (Cetak Thermal 58mm) ──────────────────────────────────────

@app.route('/slip/<int:transaksi_id>')
@login_required
def slip(transaksi_id):
    conn = get_db()
    t = conn.execute('''
        SELECT t.*, c.nama AS coa_nama, c.jenis_dana AS coa_dana,
               d.nama AS donatur_nama, u.nama AS petugas_nama
        FROM transaksi t
        LEFT JOIN chart_of_accounts c ON t.coa_id=c.id
        LEFT JOIN donatur d ON t.donatur_id=d.id
        LEFT JOIN users u ON t.user_id=u.id
        WHERE t.id=?''', (transaksi_id,)).fetchone()
    if not t:
        conn.close(); flash('Transaksi tidak ditemukan.', 'danger')
        return redirect(url_for('marketing_dashboard'))
    inst = get_instansi(conn)
    conn.close()

    jumlah = t['jumlah'] or 0
    created = t['created_at'] or ''
    try:
        tgl = datetime.strptime(created, '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y %H:%M')
    except (ValueError, TypeError):
        tgl = t['tanggal'] or ''
    dana_label = LABEL_DANA.get(t['jenis_dana'] or t['coa_dana'] or '', '')

    s = {
        'instansi':  (inst.get('nama_lembaga') or inst.get('nama') or 'BAITUL MAAL').upper(),
        'alamat':    inst.get('alamat', '') or '',
        'telepon':   inst.get('telepon', '') or '',
        'website':   inst.get('website', '') or '',
        'judul':     'BUKTI PENERIMAAN DONASI' if t['jenis'] == 'masuk' else 'BUKTI PENYALURAN',
        'no':        f"TRX-{transaksi_id:05d}",
        'tanggal':   tgl,
        'donatur':   t['donatur_nama'] or 'Umum',
        'dana':      t['coa_nama'] or dana_label or '-',
        'metode':    'Tunai',
        'jumlah_fmt': format_rupiah(jumlah),
        'terbilang': (terbilang(jumlah).capitalize() + ' rupiah'),
        'petugas':   t['petugas_nama'] or '-',
        'jenis':     t['jenis'],
    }
    return render_template('marketing/slip.html', s=s)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
